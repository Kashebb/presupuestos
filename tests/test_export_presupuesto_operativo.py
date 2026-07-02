import asyncio
import io
import sys
import unittest
from pathlib import Path

from fastapi import HTTPException
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from app.models.base import Base
from app.models.apu import APU, APUItem
from app.models.presupuesto import NodoPresupuesto, Proyecto
from app.models.recurso import Recurso
from app.api.apus import calcular_costo_apu
from app.api.presupuestos import (
    CambiarVarianteAPURequest,
    CrearVarianteAPURequest,
    cambiar_variante_apu,
    crear_variante_apu,
    exportar_presupuesto_operativo,
)


async def _read_response(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


class ExportPresupuestoOperativoTest(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
        Base.metadata.create_all(bind=engine)
        self.Session = sessionmaker(bind=engine)
        self.db = self.Session()

    def tearDown(self):
        self.db.close()

    def test_calcular_costo_apu_redondea_a_cuatro_decimales(self):
        recurso = Recurso(
            codigo="MAT-4D",
            descripcion="Material cuatro decimales",
            categoria="material",
            unidad="u",
            precio_unitario=1.23456,
        )
        apu = APU(codigo="APU-4D", nombre="APU cuatro decimales", unidad="m2", rendimiento=1.0)
        apu.items.append(APUItem(recurso=recurso, categoria="material", cantidad=2.0, orden=1))

        costo = calcular_costo_apu(apu)

        self.assertEqual(costo["precio_unitario"], 2.4692)
        self.assertEqual(costo["subtotales"]["material"], 2.4692)
        self.assertEqual(costo["herramienta_menor"], 0.0)

    def test_exporta_excel_operativo_lectura(self):
        proyecto = Proyecto(nombre="Proyecto Prueba", codigo="PR-01")
        recurso = Recurso(
            codigo="MAT-001",
            descripcion="Material prueba",
            categoria="material",
            unidad="u",
            precio_unitario=12.345,
        )
        apu = APU(codigo="APU-001", nombre="APU prueba", unidad="m2", rendimiento=1.0)
        apu.items.append(APUItem(recurso=recurso, categoria="material", cantidad=2.0, orden=1))
        self.db.add_all([proyecto, recurso, apu])
        self.db.flush()

        fase = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="FASE",
            nivel=0,
            item="1",
            descripcion="Fase principal",
            orden=1,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        rubro = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=None,
            tipo="RUBRO",
            nivel=1,
            item="1.01",
            descripcion="Rubro exportable",
            orden=2,
            unidad="m2",
            metrado=3.0,
            precio_unitario_ref=10.0,
            apu_id=apu.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
            estado_actualizacion="activo",
        )
        obsoleto = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=1,
            item="1.02",
            descripcion="Rubro obsoleto",
            orden=3,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            activo_como_rubro=True,
            estado_actualizacion="obsoleto",
        )
        self.db.add_all([fase, rubro, obsoleto])
        self.db.commit()
        rubro.padre_id = fase.id
        self.db.commit()

        response = exportar_presupuesto_operativo(proyecto.id, self.db)
        content = asyncio.run(_read_response(response))

        self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("presupuesto_operativo_pr-01.xlsx", response.headers["content-disposition"])

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        self.assertEqual(wb.sheetnames, ["Presupuesto operativo", "Resumen", "Desglose Rubros"])
        ws = wb["Presupuesto operativo"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[:5], ["Nivel", "Tipo", "Item", "Descripcion", "Unidad"])
        values_by_description = {row[3]: row for row in ws.iter_rows(min_row=2, values_only=True)}
        self.assertIn("Rubro exportable", values_by_description)
        self.assertNotIn("Rubro obsoleto", values_by_description)
        rubro_row = values_by_description["Rubro exportable"]
        self.assertEqual(rubro_row[14], "Vinculado")
        self.assertEqual(rubro_row[8], "APU-001")
        self.assertAlmostEqual(rubro_row[7], 30.0)
        self.assertAlmostEqual(rubro_row[10], 24.69)
        self.assertEqual(ws["F2"].number_format, "#,##0.0000")
        self.assertEqual(ws["G2"].number_format, "$#,##0.0000")
        detalle = wb["Desglose Rubros"]
        detalle_headers = [cell.value for cell in detalle[1]]
        self.assertEqual(detalle_headers[:5], ["Nivel", "Tipo", "Item", "Descripcion", "Unidad"])
        detalle_rows = {row[3]: row for row in detalle.iter_rows(min_row=2, values_only=True)}
        self.assertIn("Rubro exportable", detalle_rows)
        self.assertEqual(detalle_rows["Rubro exportable"][6], "APU-001")
        self.assertAlmostEqual(detalle_rows["Rubro exportable"][8], 74.07)
        self.assertAlmostEqual(detalle_rows["Rubro exportable"][13], 74.07)

    def test_exportar_proyecto_inexistente_devuelve_404(self):
        with self.assertRaises(HTTPException) as ctx:
            exportar_presupuesto_operativo(999, self.db)
        self.assertEqual(ctx.exception.status_code, 404)

    def test_crear_variante_apu_copia_items_y_asigna_rubro(self):
        proyecto = Proyecto(nombre="Proyecto Variantes", codigo="PV-01")
        recurso = Recurso(
            codigo="MAT-VAR",
            descripcion="Material variante",
            categoria="material",
            unidad="u",
            precio_unitario=5.0,
        )
        base = APU(codigo="APU-BASE", nombre="Trazado", unidad="m2", rendimiento=1.0)
        base.items.append(APUItem(recurso=recurso, categoria="material", cantidad=2.0, orden=1))
        self.db.add_all([proyecto, recurso, base])
        self.db.flush()
        rubro = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=1,
            descripcion="Trazado garita",
            orden=1,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            apu_id=base.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
        )
        self.db.add(rubro)
        self.db.commit()

        result = crear_variante_apu(
            rubro.id,
            CrearVarianteAPURequest(variante_nombre="GARITA", copiar_desde_apu_id=base.id),
            self.db,
        )

        variante = self.db.query(APU).filter(APU.id == result["apu_id"]).first()
        self.assertTrue(result["created"])
        self.assertEqual(variante.codigo, "APU-BASE-V001")
        self.assertEqual(variante.variante_nombre, "GARITA")
        self.assertEqual(variante.apu_base_id, base.id)
        self.assertEqual(variante.proyecto_id, proyecto.id)
        self.assertEqual(rubro.apu_id, variante.id)
        self.assertEqual(len(variante.items), 1)
        self.assertEqual(variante.items[0].recurso_id, recurso.id)

    def test_crear_variante_nombre_existente_asigna_sin_duplicar(self):
        proyecto = Proyecto(nombre="Proyecto Duplicados", codigo="PV-02")
        base = APU(codigo="APU-DUP", nombre="Trazado", unidad="m2", rendimiento=1.0)
        variante = APU(
            codigo="APU-DUP-V001",
            nombre="Trazado - GARITA",
            unidad="m2",
            rendimiento=1.0,
            es_variante=True,
            variante_nombre="GARITA",
        )
        self.db.add_all([proyecto, base])
        self.db.flush()
        variante.apu_base_id = base.id
        variante.proyecto_id = proyecto.id
        self.db.add(variante)
        self.db.flush()
        rubro = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=1,
            descripcion="Trazado garita",
            orden=1,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            apu_id=base.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
        )
        self.db.add(rubro)
        self.db.commit()

        result = crear_variante_apu(
            rubro.id,
            CrearVarianteAPURequest(variante_nombre=" garita ", copiar_desde_apu_id=base.id),
            self.db,
        )

        variantes = self.db.query(APU).filter(APU.es_variante.is_(True)).all()
        self.assertFalse(result["created"])
        self.assertEqual(len(variantes), 1)
        self.assertEqual(rubro.apu_id, variante.id)

    def test_cambiar_variante_rechaza_otro_apu_base(self):
        proyecto = Proyecto(nombre="Proyecto Incompatible", codigo="PV-03")
        base_a = APU(codigo="APU-A", nombre="A", unidad="m2", rendimiento=1.0)
        base_b = APU(codigo="APU-B", nombre="B", unidad="m2", rendimiento=1.0)
        self.db.add_all([proyecto, base_a, base_b])
        self.db.flush()
        variante_b = APU(
            codigo="APU-B-V001",
            nombre="B - GARITA",
            unidad="m2",
            rendimiento=1.0,
            es_variante=True,
            apu_base_id=base_b.id,
            proyecto_id=proyecto.id,
            variante_nombre="GARITA",
        )
        rubro = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=1,
            descripcion="Rubro A",
            orden=1,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            apu_id=base_a.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
        )
        self.db.add_all([variante_b, rubro])
        self.db.commit()

        with self.assertRaises(HTTPException) as ctx:
            cambiar_variante_apu(rubro.id, CambiarVarianteAPURequest(variante_apu_id=variante_b.id), self.db)
        self.assertEqual(ctx.exception.status_code, 400)

    def test_exporta_con_apu_variante_efectivo(self):
        proyecto = Proyecto(nombre="Proyecto Export Variante", codigo="PV-04")
        recurso = Recurso(
            codigo="MAT-EXP",
            descripcion="Material export",
            categoria="material",
            unidad="u",
            precio_unitario=5.0,
        )
        base = APU(codigo="APU-EXP", nombre="Trazado", unidad="m2", rendimiento=1.0)
        variante = APU(
            codigo="APU-EXP-V001",
            nombre="Trazado - GARITA",
            unidad="m2",
            rendimiento=1.0,
            es_variante=True,
            variante_nombre="GARITA",
        )
        variante.items.append(APUItem(recurso=recurso, categoria="material", cantidad=3.0, orden=1))
        self.db.add_all([proyecto, recurso, base, variante])
        self.db.flush()
        variante.apu_base_id = base.id
        variante.proyecto_id = proyecto.id
        rubro = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=1,
            item="1.01",
            descripcion="Rubro variante",
            orden=1,
            unidad="m2",
            metrado=2.0,
            precio_unitario_ref=1.0,
            apu_id=variante.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
            estado_actualizacion="activo",
        )
        self.db.add(rubro)
        self.db.commit()

        response = exportar_presupuesto_operativo(proyecto.id, self.db)
        content = asyncio.run(_read_response(response))

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Presupuesto operativo"]
        values_by_description = {row[3]: row for row in ws.iter_rows(min_row=2, values_only=True)}
        rubro_row = values_by_description["Rubro variante"]
        self.assertEqual(rubro_row[8], "APU-EXP-V001")
        self.assertEqual(rubro_row[9], "Trazado - GARITA")
        self.assertAlmostEqual(rubro_row[10], 15.0)

    def test_exporta_solo_rama_seleccionada(self):
        proyecto = Proyecto(nombre="Proyecto Rama", codigo="PV-05")
        recurso = Recurso(
            codigo="MAT-RAMA",
            descripcion="Material rama",
            categoria="material",
            unidad="u",
            precio_unitario=2.0,
        )
        apu = APU(codigo="APU-RAMA", nombre="APU rama", unidad="m2", rendimiento=1.0)
        apu.items.append(APUItem(recurso=recurso, categoria="material", cantidad=1.0, orden=1))
        self.db.add_all([proyecto, recurso, apu])
        self.db.flush()

        grupo_a = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="CAPITULO",
            nivel=0,
            item="1",
            descripcion="Grupo A",
            orden=1,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        grupo_b = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="CAPITULO",
            nivel=0,
            item="2",
            descripcion="Grupo B",
            orden=2,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        self.db.add_all([grupo_a, grupo_b])
        self.db.flush()
        rubro_a = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=grupo_a.id,
            tipo="RUBRO",
            nivel=1,
            item="1.01",
            descripcion="Rubro A",
            orden=3,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            apu_id=apu.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
            estado_actualizacion="activo",
        )
        rubro_b = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=grupo_b.id,
            tipo="RUBRO",
            nivel=1,
            item="2.01",
            descripcion="Rubro B",
            orden=4,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            apu_id=apu.id,
            activo_como_rubro=True,
            tipo_rubro="VINCULADO",
            estado_actualizacion="activo",
        )
        self.db.add_all([rubro_a, rubro_b])
        self.db.commit()

        response = exportar_presupuesto_operativo(proyecto.id, self.db, grupo_a.id)
        content = asyncio.run(_read_response(response))

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Presupuesto operativo"]
        descriptions = [row[3] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn("Grupo A", descriptions)
        self.assertIn("Rubro A", descriptions)
        self.assertNotIn("Grupo B", descriptions)
        self.assertNotIn("Rubro B", descriptions)

    def test_exporta_item_calculado_desde_jerarquia_real(self):
        proyecto = Proyecto(nombre="Proyecto Jerarquia", codigo="PJ-01")
        self.db.add(proyecto)
        self.db.flush()

        acabados = NodoPresupuesto(
            proyecto_id=proyecto.id,
            tipo="RUBRO",
            nivel=0,
            item="01.01.01.07.01.02.27.28.04",
            descripcion="ACABADOS",
            orden=1,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        self.db.add(acabados)
        self.db.flush()

        mamposteria = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=acabados.id,
            tipo="RUBRO",
            nivel=1,
            item="01.01.01.07.01.02.27.28.04.01",
            descripcion="MAMPOSTERIA",
            orden=2,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        gypsum = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=acabados.id,
            tipo="RUBRO",
            nivel=1,
            item="01.01.01.07.01.02.27.28.04.05",
            descripcion="GYPSUM Y PINTURA",
            orden=4,
            activo_como_rubro=False,
            estado_actualizacion="activo",
        )
        self.db.add_all([mamposteria, gypsum])
        self.db.flush()

        pared = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=mamposteria.id,
            tipo="RUBRO",
            nivel=2,
            item="01.01.01.07.01.02.27.28.04.02",
            descripcion="PARED DE BLOQUE",
            orden=3,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            activo_como_rubro=True,
            estado_actualizacion="activo",
        )
        cielo_raso = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=gypsum.id,
            tipo="RUBRO",
            nivel=2,
            item="01.01.01.07.01.02.27.28.04.06",
            descripcion="CIELO RASO GYPSUM",
            orden=5,
            unidad="m2",
            metrado=1.0,
            precio_unitario_ref=1.0,
            activo_como_rubro=True,
            estado_actualizacion="activo",
        )
        self.db.add_all([pared, cielo_raso])
        self.db.commit()

        response = exportar_presupuesto_operativo(proyecto.id, self.db)
        content = asyncio.run(_read_response(response))

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Presupuesto operativo"]
        items_by_description = {row[3]: row[2] for row in ws.iter_rows(min_row=2, values_only=True)}

        self.assertEqual(items_by_description["ACABADOS"], "01.01.01.07.01.02.27.28.04")
        self.assertEqual(items_by_description["MAMPOSTERIA"], "01.01.01.07.01.02.27.28.04.01")
        self.assertEqual(items_by_description["PARED DE BLOQUE"], "01.01.01.07.01.02.27.28.04.01.01")
        self.assertEqual(items_by_description["GYPSUM Y PINTURA"], "01.01.01.07.01.02.27.28.04.02")
        self.assertEqual(items_by_description["CIELO RASO GYPSUM"], "01.01.01.07.01.02.27.28.04.02.01")

        desglose = wb["Desglose Rubros"]
        desglose_items = {row[3]: row[2] for row in desglose.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(desglose_items["GYPSUM Y PINTURA"], "01.01.01.07.01.02.27.28.04.02")
        self.assertEqual(desglose_items["CIELO RASO GYPSUM"], "01.01.01.07.01.02.27.28.04.02.01")


if __name__ == "__main__":
    unittest.main()
