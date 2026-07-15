import io
import asyncio
import sys
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import HTTPException
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.models import Base
from app.models.presupuesto import Proyecto
from app.models.subcontrato import Subcontrato, SubcontratoRubro, SubcontratoRubroRecursoSnapshot
from app.services.subcontratos_excel import construir_libro_subcontrato, consolidar_materiales, libro_a_bytes, nombre_archivo_subcontrato
from app.api.subcontratos import exportar_excel


class SubcontratosExportacionTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://")
        event.listen(self.engine, "connect", lambda conn, _: conn.execute("PRAGMA foreign_keys=ON"))
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        self.db = self.Session()
        self.proyecto = Proyecto(nombre="Proyecto histórico", codigo="P-1")
        self.db.add(self.proyecto); self.db.flush()
        self.sub = Subcontrato(proyecto_id=self.proyecto.id, codigo="SC-003", nombre="Mano de obra cisterna", contratista="Contratista Uno")
        self.db.add(self.sub); self.db.flush()
        self.rubro = SubcontratoRubro(
            subcontrato_id=self.sub.id, nodo_item_snapshot="1.01", nodo_descripcion_snapshot="Hormigón histórico",
            nodo_unidad_snapshot="m3", apu_nombre_snapshot="APU histórico", preset="MANO_OBRA_EQUIPOS",
            incluye_materiales=False, incluye_mano_obra=True, incluye_equipos=True, incluye_transporte=False,
            metrado_snapshot=10.1234, pu_materiales_snapshot=11.1111, pu_mano_obra_snapshot=22.2222,
            pu_herramientas_snapshot=1.1111, pu_equipos_snapshot=3.3333, pu_transporte_snapshot=4.4444,
            pu_seleccionado_snapshot=26.6666, total_snapshot=269.3323, firma_calculo="a" * 64,
            estado_revision="DESACTUALIZADO",
        )
        self.db.add(self.rubro); self.db.flush()
        self.db.add_all([
            SubcontratoRubroRecursoSnapshot(subcontrato_rubro_id=self.rubro.id, recurso_id=None, recurso_codigo_snapshot="MAT-1", recurso_descripcion_snapshot="Cemento gris", recurso_unidad_snapshot="kg", recurso_categoria_snapshot="material", cantidad_unitaria_snapshot=2, metrado_snapshot=10.1234, cantidad_total_snapshot=20.2468, incluido_subcontrato=False),
            SubcontratoRubroRecursoSnapshot(subcontrato_rubro_id=self.rubro.id, recurso_id=None, recurso_codigo_snapshot="mat-1", recurso_descripcion_snapshot=" Cemento   gris ", recurso_unidad_snapshot="kg", recurso_categoria_snapshot="material", cantidad_unitaria_snapshot=1, metrado_snapshot=10.1234, cantidad_total_snapshot=10.1234, incluido_subcontrato=False),
            SubcontratoRubroRecursoSnapshot(subcontrato_rubro_id=self.rubro.id, recurso_id=None, recurso_codigo_snapshot="MAT-1", recurso_descripcion_snapshot="Cemento gris", recurso_unidad_snapshot="saco", recurso_categoria_snapshot="material", cantidad_unitaria_snapshot=1, metrado_snapshot=1, cantidad_total_snapshot=1, incluido_subcontrato=False),
        ])
        self.db.commit(); self.db.refresh(self.sub); self.db.refresh(self.rubro)

    def tearDown(self):
        self.db.close(); self.engine.dispose()

    def libro(self):
        contenido = libro_a_bytes(construir_libro_subcontrato(self.sub, self.proyecto, [self.rubro], fecha_exportacion=datetime(2026, 7, 15, 12)))
        return openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)

    def test_estructura_exacta_formato_y_nombre_seguro(self):
        wb = self.libro()
        self.assertEqual(wb.sheetnames, ["Subcontrato", "Desglose incluido", "Resumen"])
        ws = wb["Subcontrato"]
        self.assertEqual(ws["A1"].value, "COTIZACIÓN DE SUBCONTRATO")
        self.assertIn("A1:H1", {str(r) for r in ws.merged_cells.ranges})
        self.assertEqual(ws.page_setup.orientation, "portrait")
        self.assertEqual(ws.freeze_panes, "A12")
        self.assertGreater(ws.column_dimensions["B"].width, 30)
        self.assertEqual(nombre_archivo_subcontrato("SC-003", "Mano de obra cisterna"), "SC-003_Mano_de_obra_cisterna.xlsx")

    def test_snapshot_es_fuente_de_verdad_y_desglose_respeta_seleccion(self):
        wb = self.libro(); ws = wb["Subcontrato"]
        fila = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "1.01")
        self.assertEqual(ws.cell(fila, 2).value, "Hormigón histórico")
        self.assertEqual(ws.cell(fila, 4).value, 10.1234)
        self.assertEqual(ws.cell(fila, 6).value, 26.6666)
        self.assertEqual(ws.cell(fila, 7).value, 269.3323)
        categorias = [wb["Desglose incluido"].cell(r, 5).value for r in range(4, wb["Desglose incluido"].max_row + 1)]
        self.assertEqual(categorias, ["Mano de obra", "Herramientas menores", "Equipos"])
        self.assertNotIn("Materiales", categorias); self.assertNotIn("Transporte", categorias); self.assertNotIn("otros", categorias)

    def test_total_materiales_y_consolidacion_fallback(self):
        materiales = consolidar_materiales([self.rubro])
        self.assertEqual(len(materiales), 2)
        kg = next(m for m in materiales if m["unidad"] == "kg")
        self.assertAlmostEqual(kg["cantidad"], 30.3702, places=4)
        wb = self.libro(); ws = wb["Subcontrato"]
        total = next(ws.cell(r, 7).value for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "TOTAL SIN IVA")
        self.assertEqual(total, 269.3323)

    def test_marcas_de_estado_y_advertencia_confirmado(self):
        self.sub.estado = "BORRADOR"; self.assertIn("BORRADOR — NO APROBADO", [c.value for row in self.libro()["Subcontrato"] for c in row])
        self.sub.estado = "ANULADO"; self.assertIn("ANULADO — DOCUMENTO HISTÓRICO", [c.value for row in self.libro()["Subcontrato"] for c in row])
        self.sub.estado = "CONFIRMADO"; valores = [c.value for row in self.libro()["Subcontrato"] for c in row]
        self.assertTrue(any(isinstance(v, str) and "último estado confirmado" in v for v in valores))

    def test_sin_materiales_excluidos(self):
        for recurso in self.rubro.recursos_snapshot: recurso.incluido_subcontrato = True
        self.db.flush()
        valores = [c.value for row in self.libro()["Subcontrato"] for c in row]
        self.assertIn("No existen materiales a suministrar por la contratante.", valores)

    def test_endpoint_devuelve_xlsx_y_404(self):
        respuesta = exportar_excel(self.sub.id, self.db)
        contenido = asyncio.run(self._leer_stream(respuesta))
        self.assertEqual(respuesta.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("SC-003_Mano_de_obra_cisterna.xlsx", respuesta.headers["content-disposition"])
        self.assertEqual(openpyxl.load_workbook(io.BytesIO(contenido)).sheetnames, ["Subcontrato", "Desglose incluido", "Resumen"])
        with self.assertRaises(HTTPException) as error: exportar_excel(999999, self.db)
        self.assertEqual(error.exception.status_code, 404)

    @staticmethod
    async def _leer_stream(respuesta):
        partes = []
        async for parte in respuesta.body_iterator: partes.append(parte)
        return b"".join(partes)


if __name__ == "__main__":
    unittest.main()
