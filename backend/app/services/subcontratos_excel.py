"""Exportación histórica de subcontratos basada exclusivamente en snapshots."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


AZUL = "17365D"
AZUL_CLARO = "D9EAF7"
GRIS = "E7E6E6"
BLANCO = "FFFFFF"
ROJO = "C00000"
AMARILLO = "FFF2CC"
BORDE = Border(*( [Side(style="thin", color="A6A6A6")] * 4))
PRESETS = {
    "COMPLETO": "Completo",
    "SOLO_MATERIALES": "Solo materiales",
    "SOLO_MANO_OBRA": "Solo mano de obra",
    "MANO_OBRA_EQUIPOS": "Mano de obra + equipos",
    "MATERIALES_TRANSPORTE": "Materiales + transporte",
    "PERSONALIZADO": "Personalizado",
}
COMPONENTES = (
    ("Materiales", "incluye_materiales", "pu_materiales_snapshot"),
    ("Mano de obra", "incluye_mano_obra", "pu_mano_obra_snapshot"),
    ("Herramientas menores", "incluye_mano_obra", "pu_herramientas_snapshot"),
    ("Equipos", "incluye_equipos", "pu_equipos_snapshot"),
    ("Transporte", "incluye_transporte", "pu_transporte_snapshot"),
)


def nombre_archivo_subcontrato(codigo: str, nombre: str) -> str:
    texto = unicodedata.normalize("NFKD", nombre or "subcontrato").encode("ascii", "ignore").decode()
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", texto).strip("_") or "subcontrato"
    codigo_seguro = re.sub(r"[^A-Za-z0-9_-]+", "_", codigo or "SC").strip("_") or "SC"
    return f"{codigo_seguro}_{texto}.xlsx"


def consolidar_materiales(rubros) -> list[dict]:
    grupos = {}
    for rubro in rubros:
        for recurso in rubro.recursos_snapshot:
            if recurso.recurso_categoria_snapshot != "material" or recurso.incluido_subcontrato:
                continue
            unidad = recurso.recurso_unidad_snapshot or ""
            if recurso.recurso_id is not None:
                clave = ("id", recurso.recurso_id, unidad.casefold())
            else:
                descripcion = " ".join(recurso.recurso_descripcion_snapshot.casefold().split())
                clave = ("texto", (recurso.recurso_codigo_snapshot or "").strip().casefold(), descripcion, unidad.casefold())
            grupo = grupos.setdefault(clave, {
                "codigo": recurso.recurso_codigo_snapshot,
                "descripcion": recurso.recurso_descripcion_snapshot,
                "unidad": recurso.recurso_unidad_snapshot,
                "cantidad": 0.0,
            })
            grupo["cantidad"] += float(recurso.cantidad_total_snapshot or 0)
    return sorted(grupos.values(), key=lambda x: ((x["codigo"] or ""), x["descripcion"], x["unidad"] or ""))


def _titulo(ws, fila: int, texto: str, fin: int = 8, *, color=AZUL):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=fin)
    celda = ws.cell(fila, 1, texto)
    celda.fill = PatternFill("solid", fgColor=color)
    celda.font = Font(color=BLANCO, bold=True, size=12)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 22


def _encabezados(ws, fila: int, valores: list[str]):
    for columna, valor in enumerate(valores, 1):
        celda = ws.cell(fila, columna, valor)
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.font = Font(color=BLANCO, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE


def _estilizar_tabla(ws, inicio: int, fin: int, columnas: int):
    for fila in ws.iter_rows(min_row=inicio, max_row=fin, min_col=1, max_col=columnas):
        for celda in fila:
            celda.border = BORDE
            celda.alignment = Alignment(vertical="top", wrap_text=True)


def _configurar_hoja(ws, anchos: list[float], repetir: str | None = None):
    ws.sheet_view.showGridLines = False
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.sheet_view.zoomScale = 85
    if repetir:
        ws.print_title_rows = repetir


def construir_libro_subcontrato(subcontrato, proyecto, rubros, *, fecha_exportacion: datetime | None = None) -> Workbook:
    fecha_exportacion = fecha_exportacion or datetime.now()
    rubros = sorted(rubros, key=lambda r: ((r.nodo_item_snapshot or ""), r.id))
    materiales = consolidar_materiales(rubros)
    wb = Workbook()
    principal = wb.active
    principal.title = "Subcontrato"
    desglose = wb.create_sheet("Desglose incluido")
    resumen = wb.create_sheet("Resumen")
    _hoja_principal(principal, subcontrato, proyecto, rubros, materiales, fecha_exportacion)
    _hoja_desglose(desglose, subcontrato, rubros)
    _hoja_resumen(resumen, subcontrato, proyecto, rubros, materiales, fecha_exportacion)
    wb.calculation.fullCalcOnLoad = False
    wb.calculation.forceFullCalc = False
    return wb


def libro_a_bytes(wb: Workbook) -> bytes:
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _hoja_principal(ws, sub, proyecto, rubros, materiales, fecha):
    _configurar_hoja(ws, [20, 33, 12, 14, 25, 16, 17, 25], "9:9")
    _titulo(ws, 1, "COTIZACIÓN DE SUBCONTRATO")
    datos = [
        ("Proyecto", proyecto.nombre), ("Código", sub.codigo), ("Subcontrato", sub.nombre),
        ("Contratista", sub.contratista or "—"), ("Descripción", sub.descripcion or "—"),
        ("Fecha de exportación", fecha), ("Estado", sub.estado),
    ]
    for fila, (etiqueta, valor) in enumerate(datos, 2):
        ws.cell(fila, 1, etiqueta).font = Font(bold=True, color=AZUL)
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=8)
        ws.cell(fila, 2, valor)
        ws.cell(fila, 2).alignment = Alignment(wrap_text=True)
    ws["B7"].number_format = "yyyy-mm-dd hh:mm"
    if sub.estado in {"BORRADOR", "ANULADO"}:
        marca = "BORRADOR — NO APROBADO" if sub.estado == "BORRADOR" else "ANULADO — DOCUMENTO HISTÓRICO"
        ws.merge_cells("A9:H9")
        ws["A9"] = marca
        ws["A9"].fill = PatternFill("solid", fgColor="F4CCCC")
        ws["A9"].font = Font(color=ROJO, bold=True, size=14)
        ws["A9"].alignment = Alignment(horizontal="center")
        tabla_fila = 11
    else:
        tabla_fila = 10
        if any(r.estado_revision != "ACTUALIZADO" for r in rubros):
            ws.merge_cells("A9:H9")
            ws["A9"] = "La información exportada corresponde al último estado confirmado. Existen cambios posteriores en el presupuesto o los APUs."
            ws["A9"].fill = PatternFill("solid", fgColor=AMARILLO)
            ws["A9"].alignment = Alignment(wrap_text=True)
    _encabezados(ws, tabla_fila, ["Rubro o ítem", "Descripción", "Unidad", "Cantidad", "Configuración", "PU seleccionado", "P. Total", "Observación"])
    for fila, rubro in enumerate(rubros, tabla_fila + 1):
        valores = [rubro.nodo_item_snapshot or "—", rubro.nodo_descripcion_snapshot, rubro.nodo_unidad_snapshot or "—",
                   float(rubro.metrado_snapshot), PRESETS.get(rubro.preset, rubro.preset), float(rubro.pu_seleccionado_snapshot),
                   float(rubro.total_snapshot), f"{PRESETS.get(rubro.preset, rubro.preset)} · {rubro.estado_revision.replace('_', ' ').title()}"]
        for columna, valor in enumerate(valores, 1): ws.cell(fila, columna, valor)
    fin = tabla_fila + max(len(rubros), 1)
    if not rubros: ws.cell(tabla_fila + 1, 1, "Sin rubros asignados"); ws.merge_cells(start_row=tabla_fila + 1, start_column=1, end_row=tabla_fila + 1, end_column=8)
    _estilizar_tabla(ws, tabla_fila, fin, 8)
    for fila in range(tabla_fila + 1, fin + 1):
        ws.cell(fila, 4).number_format = "#,##0.0000"
        ws.cell(fila, 6).number_format = '"$"#,##0.0000'
        ws.cell(fila, 7).number_format = '"$"#,##0.00'
    total_fila = fin + 2
    ws.merge_cells(start_row=total_fila, start_column=1, end_row=total_fila, end_column=6)
    ws.cell(total_fila, 1, "TOTAL SIN IVA").fill = PatternFill("solid", fgColor=AZUL)
    ws.cell(total_fila, 1).font = Font(color=BLANCO, bold=True)
    ws.cell(total_fila, 7, round(sum(float(r.total_snapshot or 0) for r in rubros), 4))
    ws.cell(total_fila, 7).number_format = '"$"#,##0.00'
    ws.cell(total_fila, 7).font = Font(bold=True)
    _titulo(ws, total_fila + 3, "MATERIALES A SUMINISTRAR POR LA CONTRATANTE", 8)
    mat_header = total_fila + 4
    _encabezados(ws, mat_header, ["Código", "Descripción", "Unidad", "Cantidad"])
    ws.merge_cells(start_row=mat_header, start_column=4, end_row=mat_header, end_column=8)
    if materiales:
        for fila, material in enumerate(materiales, mat_header + 1):
            ws.cell(fila, 1, material["codigo"] or "—"); ws.cell(fila, 2, material["descripcion"]); ws.cell(fila, 3, material["unidad"] or "—")
            ws.merge_cells(start_row=fila, start_column=4, end_row=fila, end_column=8)
            ws.cell(fila, 4, round(material["cantidad"], 4)).number_format = "#,##0.0000"
        mat_fin = mat_header + len(materiales)
    else:
        mat_fin = mat_header + 1
        ws.merge_cells(start_row=mat_fin, start_column=1, end_row=mat_fin, end_column=8)
        ws.cell(mat_fin, 1, "No existen materiales a suministrar por la contratante.")
    _estilizar_tabla(ws, mat_header, mat_fin, 8)
    ws.freeze_panes = f"A{tabla_fila + 1}"
    ws.auto_filter.ref = f"A{tabla_fila}:H{fin}"
    ws.print_area = f"A1:H{mat_fin}"


def _hoja_desglose(ws, sub, rubros):
    _configurar_hoja(ws, [15, 38, 12, 14, 24, 17, 18], "3:3")
    _titulo(ws, 1, f"DESGLOSE INCLUIDO — {sub.codigo}", 7)
    _encabezados(ws, 3, ["Rubro", "Descripción", "Unidad", "Metrado", "Categoría incluida", "PU categoría", "Total categoría"])
    fila = 4
    for rubro in rubros:
        for nombre, bandera, campo in COMPONENTES:
            if not bool(getattr(rubro, bandera)):
                continue
            pu = float(getattr(rubro, campo) or 0)
            valores = [rubro.nodo_item_snapshot or "—", rubro.nodo_descripcion_snapshot, rubro.nodo_unidad_snapshot or "—",
                       float(rubro.metrado_snapshot), nombre, pu, round(float(rubro.metrado_snapshot) * pu, 4)]
            for columna, valor in enumerate(valores, 1): ws.cell(fila, columna, valor)
            fila += 1
    if fila == 4: ws.cell(4, 1, "Sin categorías incluidas"); ws.merge_cells("A4:G4"); fila = 5
    _estilizar_tabla(ws, 3, fila - 1, 7)
    for n in range(4, fila):
        ws.cell(n, 4).number_format = "#,##0.0000"; ws.cell(n, 6).number_format = '"$"#,##0.0000'; ws.cell(n, 7).number_format = '"$"#,##0.00'
    ws.freeze_panes = "A4"; ws.auto_filter.ref = f"A3:G{fila - 1}"; ws.print_area = f"A1:G{fila - 1}"


def _hoja_resumen(ws, sub, proyecto, rubros, materiales, fecha):
    _configurar_hoja(ws, [29, 27, 20, 20])
    _titulo(ws, 1, "RESUMEN DEL SUBCONTRATO", 4)
    generales = [("Proyecto", proyecto.nombre), ("Código", sub.codigo), ("Nombre", sub.nombre), ("Contratista", sub.contratista or "—"),
                 ("Estado", sub.estado), ("Fecha de creación", sub.fecha_creacion), ("Fecha de confirmación", sub.fecha_confirmacion or "—"),
                 ("Fecha de anulación", sub.fecha_anulacion or "—"), ("Fecha de exportación", fecha)]
    for fila, (etiqueta, valor) in enumerate(generales, 3):
        ws.cell(fila, 1, etiqueta).font = Font(bold=True, color=AZUL); ws.cell(fila, 2, valor)
        if isinstance(valor, datetime): ws.cell(fila, 2).number_format = "yyyy-mm-dd hh:mm"
    fila = 13; _titulo(ws, fila, "INDICADORES", 4)
    estados = defaultdict(int)
    for r in rubros: estados[r.estado_revision] += 1
    indicadores = [("Número de rubros", len(rubros)), ("Total del subcontrato", round(sum(float(r.total_snapshot or 0) for r in rubros), 4)),
                   ("Materiales a suministrar", len(materiales)), ("Rubros actualizados", estados["ACTUALIZADO"]),
                   ("Desactualizados", estados["DESACTUALIZADO"]), ("Pendientes de revisión", estados["PENDIENTE_REVISION"]), ("Errores", estados["ERROR"])]
    for n, (etiqueta, valor) in enumerate(indicadores, fila + 1): ws.cell(n, 1, etiqueta); ws.cell(n, 2, valor)
    ws.cell(fila + 2, 2).number_format = '"$"#,##0.00'
    fila = fila + len(indicadores) + 2; _titulo(ws, fila, "DISTRIBUCIÓN POR PRESET", 4)
    _encabezados(ws, fila + 1, ["Preset", "Cantidad de rubros", "Total", ""])
    por_preset = defaultdict(lambda: [0, 0.0])
    for r in rubros: por_preset[r.preset][0] += 1; por_preset[r.preset][1] += float(r.total_snapshot or 0)
    for n, preset in enumerate(PRESETS, fila + 2):
        ws.cell(n, 1, PRESETS[preset]); ws.cell(n, 2, por_preset[preset][0]); ws.cell(n, 3, round(por_preset[preset][1], 4)); ws.cell(n, 3).number_format = '"$"#,##0.00'
    _estilizar_tabla(ws, fila + 1, fila + 1 + len(PRESETS), 4)
    fila += len(PRESETS) + 3; _titulo(ws, fila, "COMPONENTES INCLUIDOS", 4)
    for n, (nombre, bandera, campo) in enumerate(COMPONENTES, fila + 1):
        total = sum(float(r.metrado_snapshot) * float(getattr(r, campo) or 0) for r in rubros if bool(getattr(r, bandera)))
        ws.cell(n, 1, nombre); ws.cell(n, 2, round(total, 4)); ws.cell(n, 2).number_format = '"$"#,##0.00'
    ws.print_area = f"A1:D{fila + len(COMPONENTES)}"
