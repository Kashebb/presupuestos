"""
Endpoints: Módulo Presupuestos
Rutas:
  GET    /presupuestos/proyectos/              → listar proyectos
  POST   /presupuestos/proyectos/              → crear proyecto vacío
  GET    /presupuestos/proyectos/{id}          → detalle proyecto
  PUT    /presupuestos/proyectos/{id}          → editar proyecto
  DELETE /presupuestos/proyectos/{id}          → eliminar proyecto + nodos
  GET    /presupuestos/proyectos/{id}/nodos    → árbol completo de nodos
  POST   /presupuestos/proyectos/{id}/importar → importar Excel
  PATCH  /presupuestos/nodos/{id}/vincular-apu → vincular APU a un nodo RUBRO
  PATCH  /presupuestos/nodos/{id}/desvincular-apu → quitar APU de un nodo
"""
import io
import json
import hashlib
import math
import re
import unicodedata
from typing import Optional, Any
from datetime import datetime

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from app import backup
from app.db import get_db
from app.models.presupuesto import (
    Proyecto,
    NodoPresupuesto,
    ActualizacionPresupuestoLote,
    PaquetePresupuesto,
    NodoAPURevision,
    NodoAPURevisionItem,
)
from app.models.apu import APU, APUItem
from app.models.recurso import Recurso
from app.api.apus import calcular_costo_apu, siguiente_codigo_apu
from app.schemas.recurso import RecursoOut

router = APIRouter(prefix="/presupuestos", tags=["Presupuestos"])

# ── Orden jerárquico para asignar padre_id al importar ──────────
ORDEN_JERARQUIA = [
    "FASE",
    "CATEGORIA",
    "SUBCATEGORIA",
    "CAPITULO",
    "SUBCAPITULO",
    "GRUPO",
    "RUBRO",
]

PU_TOLERANCIA = 0.0001

EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _texto(valor: Any) -> Optional[str]:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _numero(valor: Any) -> Optional[float]:
    if isinstance(valor, (int, float)):
        return float(valor)
    return None


def _normalizar_texto(valor: Any) -> str:
    texto = _texto(valor) or ""
    texto = " ".join(texto.split()).upper()
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def _distinto_texto(a: Any, b: Any) -> bool:
    return _normalizar_texto(a) != _normalizar_texto(b)


def _distinto_numero(a: Optional[float], b: Optional[float], tolerancia: float = 0.000001) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    return abs(float(a) - float(b)) > tolerancia


def _leer_excel_presupuesto(contenido: bytes, hoja: str) -> tuple[list[dict], dict]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"No se pudo leer el archivo Excel: {e}"
        )

    if hoja not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"Hoja '{hoja}' no encontrada. Hojas disponibles: {wb.sheetnames}",
        )

    ws = wb[hoja]
    filas = []
    stack: list[dict] = []
    conteo_por_tipo: dict[str, int] = {}

    for numero_fila, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        tipo = _texto(row[0])
        if tipo is None:
            continue
        tipo = tipo.upper()
        if tipo not in ORDEN_JERARQUIA:
            continue

        nivel_actual = ORDEN_JERARQUIA.index(tipo)
        while stack and ORDEN_JERARQUIA.index(stack[-1]["tipo"]) >= nivel_actual:
            stack.pop()

        pu_raw = row[5]
        observacion_excel = _texto(row[7]) if len(row) > 7 else None
        if tipo == "RUBRO":
            metrado = _numero(row[4])
            if isinstance(pu_raw, str) and pu_raw.strip().upper() == "SIN_APU":
                precio_ref = 0.0
                observaciones = "SIN_APU"
            elif isinstance(pu_raw, (int, float)):
                precio_ref = float(pu_raw)
                observaciones = observacion_excel
            else:
                precio_ref = 0.0
                observaciones = _texto(pu_raw) or observacion_excel
            unidad = _texto(row[3])
        else:
            metrado = None
            precio_ref = None
            observaciones = observacion_excel
            unidad = None

        item = _texto(row[1])
        descripcion = _texto(row[2]) or "(sin descripcion)"
        registro = {
            "row": numero_fila,
            "tipo": tipo,
            "item": item,
            "descripcion": descripcion,
            "unidad": unidad,
            "metrado": metrado,
            "precio_unitario_ref": precio_ref,
            "total_ref_excel": _numero(row[6]),
            "observaciones_excel": observaciones,
            "path": [p.copy() for p in stack],
        }
        filas.append(registro)
        conteo_por_tipo[tipo] = conteo_por_tipo.get(tipo, 0) + 1

        if tipo != "RUBRO":
            stack.append({
                "tipo": tipo,
                "item": item,
                "descripcion": descripcion,
                "row": numero_fila,
            })

    if not filas:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron filas validas en la hoja especificada.",
        )

    return filas, {
        "total_nodos": len(filas),
        "total_rubros": conteo_por_tipo.get("RUBRO", 0),
        "por_tipo": conteo_por_tipo,
    }


def _ruta_nodo(nodo: NodoPresupuesto, por_id: dict[int, NodoPresupuesto]) -> list[NodoPresupuesto]:
    ruta = []
    actual = nodo
    while actual is not None:
        ruta.append(actual)
        actual = por_id.get(actual.padre_id)
    return list(reversed(ruta))


def _ruta_clave_actual(nodo: NodoPresupuesto, por_id: dict[int, NodoPresupuesto]) -> list[tuple[str, Optional[str]]]:
    return [(n.tipo, n.item) for n in _ruta_nodo(nodo, por_id)[:-1]]


def _ruta_clave_excel(registro: dict) -> list[tuple[str, Optional[str]]]:
    return [(n["tipo"], n["item"]) for n in registro["path"]]


def _total_ref_rubro(nodo: NodoPresupuesto) -> float:
    if nodo.tipo != "RUBRO":
        return 0.0
    if nodo.metrado is None or nodo.precio_unitario_ref is None:
        return 0.0
    return float(nodo.metrado) * float(nodo.precio_unitario_ref)


def _total_linea(metrado: Optional[float], precio_unitario: Optional[float]) -> Optional[float]:
    if metrado is None or precio_unitario is None:
        return None
    return float(metrado) * float(precio_unitario)


def _nombre_archivo_excel(proyecto: Proyecto) -> str:
    base = proyecto.codigo or proyecto.nombre or f"proyecto-{proyecto.id}"
    base = unicodedata.normalize("NFD", base)
    base = "".join(c for c in base if unicodedata.category(c) != "Mn")
    base = re.sub(r"[^A-Za-z0-9_-]+", "_", base).strip("_").lower()
    return f"presupuesto_operativo_{base or proyecto.id}.xlsx"


def _ordenar_arbol_presupuesto(nodos: list[NodoPresupuesto]) -> list[NodoPresupuesto]:
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    ids = {n.id for n in nodos}
    for nodo in nodos:
        padre_id = nodo.padre_id if nodo.padre_id in ids else None
        hijos_por_padre.setdefault(padre_id, []).append(nodo)
    for hijos in hijos_por_padre.values():
        hijos.sort(key=lambda n: (n.orden or 0, n.id))

    ordenados: list[NodoPresupuesto] = []

    def visitar(padre_id: Optional[int]) -> None:
        for hijo in hijos_por_padre.get(padre_id, []):
            ordenados.append(hijo)
            visitar(hijo.id)

    visitar(None)
    return ordenados


def _nivel_exportacion(nodo: NodoPresupuesto, por_id: dict[int, NodoPresupuesto]) -> int:
    if nodo.nivel is not None:
        return max(0, int(nodo.nivel))
    nivel = 0
    actual = nodo
    visitados = set()
    while actual.padre_id and actual.padre_id in por_id and actual.padre_id not in visitados:
        visitados.add(actual.padre_id)
        nivel += 1
        actual = por_id[actual.padre_id]
    return nivel


def _es_rubro_operativo_desde_hijos(nodo: NodoPresupuesto, hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]]) -> bool:
    activo = bool(nodo.activo_como_rubro) if nodo.activo_como_rubro is not None else nodo.tipo == "RUBRO"
    return activo and not hijos_por_padre.get(nodo.id)


def _items_exportacion_por_jerarquia(nodos: list[NodoPresupuesto]) -> dict[int, Optional[str]]:
    por_id = {n.id for n in nodos}
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for nodo in nodos:
        padre_id = nodo.padre_id if nodo.padre_id in por_id else None
        hijos_por_padre.setdefault(padre_id, []).append(nodo)
    for hijos in hijos_por_padre.values():
        hijos.sort(key=lambda n: (n.orden or 0, n.id))

    items: dict[int, Optional[str]] = {}

    def ancho_hijos(hijos: list[NodoPresupuesto]) -> int:
        anchos = []
        for hijo in hijos:
            ultimo = (hijo.item or "").split(".")[-1]
            if ultimo.isdigit():
                anchos.append(len(ultimo))
        return max([2, *anchos])

    def visitar(padre_id: Optional[int], prefijo: Optional[str]) -> None:
        hijos = hijos_por_padre.get(padre_id, [])
        ancho = ancho_hijos(hijos)
        for idx, hijo in enumerate(hijos, start=1):
            if prefijo:
                item_exportacion = f"{prefijo}.{str(idx).zfill(ancho)}"
            else:
                item_exportacion = hijo.item or str(idx).zfill(ancho)
            items[hijo.id] = item_exportacion
            visitar(hijo.id, item_exportacion)

    visitar(None, None)
    return items


def _estado_exportacion(nodo: NodoPresupuesto, costo_apu: Optional[dict], db: Optional[Session] = None) -> str:
    return _estado_revision_apu(nodo, costo_apu, db)


def _motivos_revision_apu(nodo: NodoPresupuesto, costo_apu: Optional[dict]) -> list[dict[str, str]]:
    motivos: list[dict[str, str]] = []
    if nodo.requiere_revision_apu:
        motivos.append(
            {
                "codigo": "rubro_editado",
                "descripcion": "El rubro cambio despues de vincular el APU.",
                "detalle": [
                    "La descripcion o la unidad del rubro fue editada despues de tener un APU vinculado.",
                    "La app no puede asumir automaticamente que el APU sigue siendo compatible con el rubro.",
                    "Al marcarlo como revisado, confirmas que para este rubro la descripcion, unidad y APU vinculado siguen siendo aceptables.",
                ],
            }
        )
    if costo_apu and costo_apu.get("control_costo") == "revisar_costo":
        observacion = (getattr(nodo.apu, "observacion", None) or "").strip()
        detalle = [
            "El APU vinculado esta marcado para revisar costo.",
            "Esto normalmente significa que el precio calculado del APU no coincide con el valor de control/maestro, o que la app detecto una diferencia que no debe aceptarse automaticamente.",
            "No significa necesariamente que el rubro este mal ni que no tenga precio.",
            "Al marcarlo como revisado, validas solo este rubro; no cambias ni apruebas el APU para otros rubros.",
        ]
        if observacion:
            detalle.append(f"Observacion registrada en el APU: {observacion}")
        motivos.append(
            {
                "codigo": "costo_apu_revisar",
                "descripcion": "El APU vinculado tiene una alerta de costo.",
                "detalle": detalle,
            }
        )
    return motivos


def _firma_revision_apu(nodo: NodoPresupuesto, motivos: list[dict[str, str]]) -> str:
    payload = {
        "nodo_id": nodo.id,
        "apu_id": nodo.apu_id,
        "descripcion": nodo.descripcion or "",
        "unidad": nodo.unidad or "",
        "motivos": sorted(
            [{"codigo": motivo["codigo"]} for motivo in motivos],
            key=lambda item: item["codigo"],
        ),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _revision_validada_actual(
    db: Session,
    nodo: NodoPresupuesto,
    motivos: list[dict[str, str]],
) -> Optional[NodoAPURevision]:
    if not motivos:
        return None
    firma = _firma_revision_apu(nodo, motivos)
    return (
        db.query(NodoAPURevision)
        .filter(
            NodoAPURevision.nodo_id == nodo.id,
            NodoAPURevision.estado == "validado",
            NodoAPURevision.firma_revision == firma,
            NodoAPURevision.snapshot_descripcion == (nodo.descripcion or ""),
            NodoAPURevision.snapshot_unidad == (nodo.unidad or ""),
            NodoAPURevision.snapshot_apu_id == nodo.apu_id,
        )
        .order_by(NodoAPURevision.fecha_creacion.desc(), NodoAPURevision.id.desc())
        .first()
    )


def _estado_revision_apu(
    nodo: NodoPresupuesto,
    costo_apu: Optional[dict],
    db: Optional[Session] = None,
) -> str:
    if nodo.observaciones == "SIN_APU":
        return "Subcontratado"
    if not nodo.apu_id:
        return "Pendiente"
    motivos = _motivos_revision_apu(nodo, costo_apu)
    if motivos:
        if db and _revision_validada_actual(db, nodo, motivos):
            return "Validado"
        return "Revisar"
    return "Vinculado"


def _costo_apu_con_control(apu: Optional[APU]) -> Optional[dict]:
    if not apu:
        return None
    costo = calcular_costo_apu(apu)
    costo["control_costo"] = (
        "revisar_costo"
        if "COSTO_NO_COINCIDE_CON_MAESTRO" in (apu.observacion or "")
        else "ok"
    )
    return costo


def _revision_apu_out(nodo: NodoPresupuesto, db: Session, costo_apu: Optional[dict] = None) -> RevisionAPUOut:
    costo = costo_apu if costo_apu is not None else _costo_apu_con_control(nodo.apu)
    motivos = _motivos_revision_apu(nodo, costo)
    revision_actual = _revision_validada_actual(db, nodo, motivos) if motivos else None
    estado = _estado_revision_apu(nodo, costo, db)
    items_actuales: dict[str, RevisionAPUHistorialItemOut] = {}
    if revision_actual:
        items_actuales = {
            item.codigo_motivo: RevisionAPUHistorialItemOut(
                codigo_motivo=item.codigo_motivo,
                descripcion_motivo=item.descripcion_motivo,
                aprobado=bool(item.aprobado),
                comentario=item.comentario,
            )
            for item in revision_actual.items
        }

    historial = (
        db.query(NodoAPURevision)
        .filter(NodoAPURevision.nodo_id == nodo.id)
        .order_by(NodoAPURevision.fecha_creacion.desc(), NodoAPURevision.id.desc())
        .all()
    )
    return RevisionAPUOut(
        nodo_id=nodo.id,
        estado=estado,
        motivos_actuales=[
            RevisionAPUMotivoOut(
                codigo=motivo["codigo"],
                descripcion=motivo["descripcion"],
                detalle=motivo.get("detalle") or [],
                aprobado=items_actuales.get(motivo["codigo"]).aprobado if motivo["codigo"] in items_actuales else False,
                comentario=items_actuales.get(motivo["codigo"]).comentario if motivo["codigo"] in items_actuales else None,
            )
            for motivo in motivos
        ],
        historial=[
            RevisionAPUHistorialOut(
                id=revision.id,
                estado=revision.estado,
                items=[
                    RevisionAPUHistorialItemOut(
                        codigo_motivo=item.codigo_motivo,
                        descripcion_motivo=item.descripcion_motivo,
                        aprobado=bool(item.aprobado),
                        comentario=item.comentario,
                    )
                    for item in revision.items
                ],
            )
            for revision in historial
        ],
    )


VISTAS_EXPORTACION_ESTADO = {
    "vinculados": "Vinculado",
    "pendientes": "Pendiente",
    "subcontratados": "Subcontratado",
    "revisar": "Revisar",
    "validados": "Validado",
}


def _filtrar_nodos_por_vista_exportacion(nodos: list[NodoPresupuesto], vista_exportacion: Optional[str], db: Session) -> list[NodoPresupuesto]:
    vista = (vista_exportacion or "todo").strip().lower()
    if vista in ("", "todo"):
        return nodos
    estado_objetivo = VISTAS_EXPORTACION_ESTADO.get(vista)
    if not estado_objetivo:
        raise HTTPException(status_code=400, detail="Vista de exportacion no permitida")

    nodos_activos = [n for n in nodos if (n.estado_actualizacion or "activo") != "obsoleto"]
    por_id = {n.id: n for n in nodos_activos}
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for nodo in nodos_activos:
        padre_id = nodo.padre_id if nodo.padre_id in por_id else None
        hijos_por_padre.setdefault(padre_id, []).append(nodo)

    costos_por_apu: dict[int, dict] = {}
    ids_incluir: set[int] = set()
    for nodo in nodos_activos:
        if not _es_rubro_operativo_desde_hijos(nodo, hijos_por_padre):
            continue
        costo_apu = None
        if nodo.apu:
            if nodo.apu.id not in costos_por_apu:
                costo = calcular_costo_apu(nodo.apu)
                costo["control_costo"] = (
                    "revisar_costo"
                    if "COSTO_NO_COINCIDE_CON_MAESTRO" in (nodo.apu.observacion or "")
                    else "ok"
                )
                costos_por_apu[nodo.apu.id] = costo
            costo_apu = costos_por_apu[nodo.apu.id]
        estado = _estado_exportacion(nodo, costo_apu, db)
        if estado_objetivo == "Vinculado":
            coincide = estado in ("Vinculado", "Validado")
        else:
            coincide = estado == estado_objetivo
        if not coincide:
            continue
        actual: Optional[NodoPresupuesto] = nodo
        while actual is not None:
            ids_incluir.add(actual.id)
            actual = por_id.get(actual.padre_id)

    return [n for n in nodos if n.id in ids_incluir]


def _crear_workbook_presupuesto_operativo(proyecto: Proyecto, nodos: list[NodoPresupuesto], db: Optional[Session] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto operativo"
    resumen = wb.create_sheet("Resumen")
    desglose = wb.create_sheet("Desglose Rubros")

    nodos_activos = [n for n in nodos if (n.estado_actualizacion or "activo") != "obsoleto"]
    por_id = {n.id: n for n in nodos_activos}
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for nodo in nodos_activos:
        padre_id = nodo.padre_id if nodo.padre_id in por_id else None
        hijos_por_padre.setdefault(padre_id, []).append(nodo)

    ordenados = _ordenar_arbol_presupuesto(nodos_activos)
    items_exportacion = _items_exportacion_por_jerarquia(nodos_activos)
    costos_por_apu: dict[int, dict] = {}
    acumulados: dict[int, dict[str, float]] = {n.id: {"ref": 0.0, "apu": 0.0} for n in nodos_activos}
    acumulados_desglose: dict[int, dict[str, float]] = {
        n.id: {"material": 0.0, "mano_de_obra": 0.0, "equipo": 0.0, "transporte": 0.0, "herramienta_menor": 0.0}
        for n in nodos_activos
    }
    rows_data = []
    conteos = {"estructura": 0, "rubros": 0, "Vinculado": 0, "Validado": 0, "Pendiente": 0, "Subcontratado": 0, "Revisar": 0}

    for nodo in ordenados:
        es_rubro = _es_rubro_operativo_desde_hijos(nodo, hijos_por_padre)
        nivel = _nivel_exportacion(nodo, por_id)
        costo_apu = None
        pu_apu = None
        total_apu = None
        total_ref = _total_linea(nodo.metrado, nodo.precio_unitario_ref) if es_rubro else None

        if es_rubro and nodo.apu:
            if nodo.apu.id not in costos_por_apu:
                costo = calcular_costo_apu(nodo.apu)
                costo["control_costo"] = (
                    "revisar_costo"
                    if "COSTO_NO_COINCIDE_CON_MAESTRO" in (nodo.apu.observacion or "")
                    else "ok"
                )
                costos_por_apu[nodo.apu.id] = costo
            costo_apu = costos_por_apu[nodo.apu.id]
            pu_apu = costo_apu["precio_unitario"]
            total_apu = _total_linea(nodo.metrado, pu_apu)

        estado = _estado_exportacion(nodo, costo_apu, db) if es_rubro else ""
        if es_rubro:
            conteos["rubros"] += 1
            conteos[estado] += 1
            ref_value = float(total_ref or 0.0)
            apu_value = float(total_apu or 0.0)
            metrado_value = float(nodo.metrado or 0.0)
            rubro_desglose = {"material": 0.0, "mano_de_obra": 0.0, "equipo": 0.0, "transporte": 0.0, "herramienta_menor": 0.0}
            if costo_apu:
                subtotales = costo_apu.get("subtotales") or {}
                for categoria in ("material", "mano_de_obra", "equipo", "transporte"):
                    rubro_desglose[categoria] = float(subtotales.get(categoria, 0.0) or 0.0) * metrado_value
                rubro_desglose["herramienta_menor"] = float(costo_apu.get("herramienta_menor", 0.0) or 0.0) * metrado_value
            actual = nodo
            while actual is not None:
                if actual.id in acumulados:
                    acumulados[actual.id]["ref"] += ref_value
                    acumulados[actual.id]["apu"] += apu_value
                if actual.id in acumulados_desglose:
                    for categoria, valor in rubro_desglose.items():
                        acumulados_desglose[actual.id][categoria] += valor
                actual = por_id.get(actual.padre_id)
        else:
            conteos["estructura"] += 1

        rows_data.append({
            "nodo": nodo,
            "nivel": nivel,
            "item": items_exportacion.get(nodo.id, nodo.item),
            "es_rubro": es_rubro,
            "pu_apu": pu_apu,
            "total_ref": total_ref,
            "total_apu": total_apu,
            "estado": estado,
            "costo_apu": costo_apu,
        })

    headers = [
        "Nivel", "Tipo", "Item", "Descripcion", "Unidad", "Metrado", "PU ref", "Total ref",
        "APU", "Nombre APU", "PU APU", "Total APU", "Dif $", "Dif %", "Estado", "Observaciones",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    group_fill = PatternFill("solid", fgColor="E5E7EB")
    line_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for data in rows_data:
        nodo = data["nodo"]
        es_rubro = data["es_rubro"]
        ref = data["total_ref"] if es_rubro else acumulados[nodo.id]["ref"]
        apu_total = data["total_apu"] if es_rubro else (acumulados[nodo.id]["apu"] or None)
        diff = apu_total - ref if ref and apu_total is not None else None
        diff_pct = diff / ref if ref and diff is not None else None
        ws.append([
            data["nivel"],
            nodo.tipo,
            data["item"],
            nodo.descripcion,
            nodo.unidad if es_rubro else None,
            nodo.metrado if es_rubro else None,
            nodo.precio_unitario_ref if es_rubro else None,
            ref,
            nodo.apu.codigo if es_rubro and nodo.apu else None,
            nodo.apu.nombre if es_rubro and nodo.apu else None,
            data["pu_apu"],
            apu_total,
            diff,
            diff_pct,
            data["estado"],
            nodo.observaciones,
        ])
        row = ws[ws.max_row]
        fill = line_fill if es_rubro else group_fill
        for cell in row:
            cell.fill = fill
            cell.border = border
        row[3].alignment = Alignment(indent=min(data["nivel"], 15), wrap_text=True)
        if not es_rubro:
            for cell in row:
                cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=14):
        for idx, cell in enumerate(row, start=6):
            if idx == 6:
                cell.number_format = '#,##0.0000'
            elif idx == 14:
                cell.number_format = '0.0000%'
            else:
                cell.number_format = '$#,##0.0000'

    widths = [8, 15, 14, 55, 12, 12, 14, 14, 18, 36, 14, 14, 14, 12, 16, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    resumen.append(["Campo", "Valor"])
    resumen.append(["Proyecto", proyecto.nombre])
    resumen.append(["Codigo", proyecto.codigo or ""])
    resumen.append(["Fecha generacion", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    resumen.append(["Filas estructura", conteos["estructura"]])
    resumen.append(["Rubros operativos", conteos["rubros"]])
    resumen.append(["Vinculados", conteos["Vinculado"] + conteos["Validado"]])
    resumen.append(["Validados", conteos["Validado"]])
    resumen.append(["Pendientes", conteos["Pendiente"]])
    resumen.append(["Subcontratados", conteos["Subcontratado"]])
    resumen.append(["Revisar", conteos["Revisar"]])
    total_ref = sum(acum["ref"] for nodo_id, acum in acumulados.items() if por_id[nodo_id].padre_id is None)
    total_apu = sum(acum["apu"] for nodo_id, acum in acumulados.items() if por_id[nodo_id].padre_id is None)
    resumen.append(["Total ref", total_ref])
    resumen.append(["Total APU", total_apu])
    resumen.append(["Diferencia", total_apu - total_ref])

    for cell in resumen[1]:
        cell.fill = header_fill
        cell.font = header_font
    resumen.column_dimensions["A"].width = 24
    resumen.column_dimensions["B"].width = 38
    for cell in resumen["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.0000' if cell.row >= 11 else '#,##0'

    desglose_headers = [
        "Nivel",
        "Tipo",
        "Item",
        "Descripcion",
        "Unidad",
        "Metrado",
        "APU",
        "Variante",
        "Materiales",
        "Mano de obra",
        "Equipos",
        "Transporte",
        "Herr. menor",
        "Total APU",
    ]
    desglose.append(desglose_headers)
    for cell in desglose[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for data in rows_data:
        nodo = data["nodo"]
        es_rubro = data["es_rubro"]
        apu = nodo.apu if es_rubro else None
        valores = acumulados_desglose[nodo.id]
        total_apu_desglose = data["total_apu"] if es_rubro else (acumulados[nodo.id]["apu"] or None)
        desglose.append([
            data["nivel"],
            nodo.tipo,
            data["item"],
            nodo.descripcion,
            nodo.unidad if es_rubro else None,
            nodo.metrado if es_rubro else None,
            apu.codigo if apu else None,
            apu.variante_nombre if apu and getattr(apu, "es_variante", False) else ("Base" if apu else None),
            valores["material"],
            valores["mano_de_obra"],
            valores["equipo"],
            valores["transporte"],
            valores["herramienta_menor"],
            total_apu_desglose,
        ])
        row = desglose[desglose.max_row]
        fill = line_fill if es_rubro else group_fill
        for cell in row:
            cell.fill = fill
            cell.border = border
        row[3].alignment = Alignment(indent=min(data["nivel"], 15), wrap_text=True)
        if not es_rubro:
            for cell in row:
                cell.font = Font(bold=True)

    for row in desglose.iter_rows(min_row=2, min_col=6, max_col=14):
        for idx, cell in enumerate(row, start=6):
            if idx == 6:
                cell.number_format = '#,##0.0000'
            elif idx in (7, 8):
                continue
            else:
                cell.number_format = '$#,##0.0000'
    desglose_widths = [8, 15, 14, 55, 12, 12, 18, 18, 14, 14, 14, 14, 14, 16]
    for idx, width in enumerate(desglose_widths, start=1):
        desglose.column_dimensions[get_column_letter(idx)].width = width
    desglose.freeze_panes = "A2"
    desglose.auto_filter.ref = desglose.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _total_ref_excel(registro: dict) -> float:
    if registro["tipo"] != "RUBRO":
        return 0.0
    if registro["metrado"] is None or registro["precio_unitario_ref"] is None:
        return 0.0
    return float(registro["metrado"]) * float(registro["precio_unitario_ref"])


def _detalle_rubro(nodo: NodoPresupuesto) -> dict:
    return {
        "id": nodo.id,
        "item": nodo.item,
        "descripcion": nodo.descripcion,
        "unidad": nodo.unidad,
        "metrado": nodo.metrado,
        "precio_unitario_ref": nodo.precio_unitario_ref,
        "apu_id": nodo.apu_id,
        "tipo_rubro": nodo.tipo_rubro,
        "observaciones": nodo.observaciones,
        "individualizado": nodo.individualizado,
        "estado_actualizacion": nodo.estado_actualizacion or "activo",
    }


def _detalle_excel(registro: dict) -> dict:
    return {
        "row": registro["row"],
        "item": registro["item"],
        "descripcion": registro["descripcion"],
        "unidad": registro["unidad"],
        "metrado": registro["metrado"],
        "precio_unitario_ref": registro["precio_unitario_ref"],
        "observaciones_excel": registro["observaciones_excel"],
        "ruta": " > ".join([p["descripcion"] for p in registro["path"]]),
    }


def _siguiente_item_rubro(db: Session, proyecto_id: int, padre_id: Optional[int], referencia_item: Optional[str]) -> Optional[str]:
    if referencia_item:
        partes = referencia_item.split(".")
        ultimo = partes[-1]
        if ultimo.isdigit():
            ancho = len(ultimo)
            base = partes[:-1]
            valor = int(ultimo) + 1
            existentes = {
                item for (item,) in db.query(NodoPresupuesto.item)
                .filter(NodoPresupuesto.proyecto_id == proyecto_id)
                .filter(NodoPresupuesto.item.isnot(None))
                .all()
            }
            while True:
                candidato = ".".join([*base, str(valor).zfill(ancho)])
                if candidato not in existentes:
                    return candidato
                valor += 1

    padre = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == padre_id).first() if padre_id else None
    prefijo = padre.item if padre and padre.item else None
    if not prefijo:
        return None

    existentes_hermanos = [
        item for (item,) in db.query(NodoPresupuesto.item)
        .filter(
            NodoPresupuesto.proyecto_id == proyecto_id,
            NodoPresupuesto.padre_id == padre_id,
            NodoPresupuesto.tipo == "RUBRO",
            NodoPresupuesto.item.isnot(None),
        )
        .all()
    ]
    maximo = 0
    for item in existentes_hermanos:
        if not item.startswith(f"{prefijo}."):
            continue
        ultimo = item.split(".")[-1]
        if ultimo.isdigit():
            maximo = max(maximo, int(ultimo))
    return f"{prefijo}.{str(maximo + 1).zfill(2)}"


def _renumerar_rubros_hermanos_desde(
    db: Session,
    proyecto_id: int,
    padre_id: Optional[int],
    referencia_item: Optional[str],
    orden_referencia: int,
) -> None:
    if not referencia_item:
        return
    partes = referencia_item.split(".")
    ultimo = partes[-1]
    if not ultimo.isdigit():
        return

    base = partes[:-1]
    ancho = len(ultimo)
    siguiente = int(ultimo) + 1
    rubros = (
        db.query(NodoPresupuesto)
        .filter(
            NodoPresupuesto.proyecto_id == proyecto_id,
            NodoPresupuesto.padre_id == padre_id,
            NodoPresupuesto.tipo == "RUBRO",
            NodoPresupuesto.orden > orden_referencia,
        )
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    for idx, rubro in enumerate(rubros):
        rubro.item = ".".join([*base, str(siguiente + idx).zfill(ancho)])


def _renumerar_rubros_activos_hermanos(db: Session, proyecto_id: int, padre_id: Optional[int]) -> None:
    padre = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == padre_id).first() if padre_id else None
    rubros = (
        db.query(NodoPresupuesto)
        .filter(
            NodoPresupuesto.proyecto_id == proyecto_id,
            NodoPresupuesto.padre_id == padre_id,
            NodoPresupuesto.tipo == "RUBRO",
            or_(NodoPresupuesto.estado_actualizacion != "obsoleto", NodoPresupuesto.estado_actualizacion.is_(None)),
        )
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    if not rubros:
        return

    if padre and padre.item:
        base = padre.item.split(".")
        ancho = 2
    else:
        partes = (rubros[0].item or "").split(".")
        ultimo = partes[-1] if partes else ""
        base = partes[:-1]
        ancho = len(ultimo) if ultimo.isdigit() else 2

    for idx, rubro in enumerate(rubros, start=1):
        rubro.item = ".".join([*base, str(idx).zfill(ancho)])


def _construir_preview_actualizacion(
    db: Session,
    proyecto_id: int,
    contenido: bytes,
    archivo_nombre: Optional[str],
    hoja: str,
) -> dict:
    filas_excel, stats_excel = _leer_excel_presupuesto(contenido, hoja)
    nodos_actuales = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    por_id = {n.id: n for n in nodos_actuales}
    actuales_por_item = {n.item: n for n in nodos_actuales if n.item}
    rubros_actuales = [n for n in nodos_actuales if n.tipo == "RUBRO"]
    rubros_actuales_por_item = {n.item: n for n in rubros_actuales if n.item}
    rubros_excel = [f for f in filas_excel if f["tipo"] == "RUBRO" and f["item"]]
    rubros_excel_por_item = {f["item"]: f for f in rubros_excel}

    items_actuales = set(rubros_actuales_por_item)
    items_excel = set(rubros_excel_por_item)
    items_coincidentes = items_actuales & items_excel
    items_nuevos = items_excel - items_actuales
    items_obsoletos = items_actuales - items_excel

    actualizar_seguro = []
    revision = []
    matriz = []

    for item in sorted(items_coincidentes):
        actual = rubros_actuales_por_item[item]
        nuevo = rubros_excel_por_item[item]
        cambios = {
            "descripcion": _distinto_texto(actual.descripcion, nuevo["descripcion"]),
            "unidad": _distinto_texto(actual.unidad, nuevo["unidad"]),
            "metrado": _distinto_numero(actual.metrado, nuevo["metrado"]),
            "precio_unitario_ref": _distinto_numero(
                actual.precio_unitario_ref,
                nuevo["precio_unitario_ref"],
                PU_TOLERANCIA,
            ),
            "estructura": _ruta_clave_actual(actual, por_id) != _ruta_clave_excel(nuevo),
        }
        motivos = []
        if cambios["unidad"]:
            motivos.append("CAMBIO_UNIDAD")
        if actual.apu_id and (cambios["descripcion"] or cambios["unidad"]):
            motivos.append("APU_VINCULADO_CON_CAMBIO_NOMBRE_O_UNIDAD")
        if cambios["estructura"]:
            motivos.append("CAMBIO_ESTRUCTURA")

        entrada = {
            "tipo": "coincidente",
            "item": item,
            "actual": _detalle_rubro(actual),
            "excel": _detalle_excel(nuevo),
            "cambios": cambios,
            "motivos_revision": motivos,
        }
        matriz.append(entrada)
        if motivos:
            revision.append(entrada)
        elif any(cambios.values()):
            actualizar_seguro.append(entrada)

    obsoletos = [
        {
            "tipo": "obsoleto",
            "item": item,
            "actual": _detalle_rubro(rubros_actuales_por_item[item]),
            "motivos_revision": [],
        }
        for item in sorted(items_obsoletos)
    ]

    nuevos_por_ancla: dict[str, dict] = {}
    for item in sorted(items_nuevos):
        registro = rubros_excel_por_item[item]
        ancla = None
        for ancestro in registro["path"]:
            if ancestro["item"] in actuales_por_item:
                existente = actuales_por_item[ancestro["item"]]
                ancla = {
                    "id": existente.id,
                    "tipo": existente.tipo,
                    "item": existente.item,
                    "descripcion": existente.descripcion,
                }
        clave = ancla["item"] if ancla else "__raiz__"
        if clave not in nuevos_por_ancla:
            nuevos_por_ancla[clave] = {
                "tipo": "nuevos",
                "ancla": ancla,
                "ruta": " > ".join([p["descripcion"] for p in registro["path"]]),
                "rubros": [],
            }
        nuevos_por_ancla[clave]["rubros"].append(_detalle_excel(registro))

    nuevos_paquetes = sorted(
        nuevos_por_ancla.values(),
        key=lambda p: (-len(p["rubros"]), p["ruta"]),
    )

    total_actual = sum(_total_ref_rubro(n) for n in rubros_actuales if (n.estado_actualizacion or "activo") != "obsoleto")
    total_excel = sum(_total_ref_excel(f) for f in rubros_excel)
    resumen = {
        "archivo": archivo_nombre,
        "hoja": hoja,
        "total_nodos_actual": len(nodos_actuales),
        "total_rubros_actual": len(rubros_actuales),
        "total_ref_actual": total_actual,
        "total_nodos_excel": stats_excel["total_nodos"],
        "total_rubros_excel": stats_excel["total_rubros"],
        "total_ref_excel": total_excel,
        "rubros_coincidentes": len(items_coincidentes),
        "rubros_nuevos": len(items_nuevos),
        "rubros_obsoletos": len(items_obsoletos),
        "rubros_actualizacion_segura": len(actualizar_seguro),
        "rubros_revision": len(revision),
        "cambios_nombre": sum(1 for m in matriz if m["cambios"]["descripcion"]),
        "cambios_unidad": sum(1 for m in matriz if m["cambios"]["unidad"]),
        "cambios_metrado": sum(1 for m in matriz if m["cambios"]["metrado"]),
        "cambios_pu": sum(1 for m in matriz if m["cambios"]["precio_unitario_ref"]),
        "cambios_estructura": sum(1 for m in matriz if m["cambios"]["estructura"]),
    }

    return {
        "resumen": resumen,
        "paquetes": {
            "automatico_seguro": {
                "rubros_existentes": actualizar_seguro,
                "nuevos_por_ancla": nuevos_paquetes,
            },
            "revision_requerida": revision,
            "obsoletos": obsoletos,
        },
        "matriz": matriz,
        "_filas_excel": filas_excel,
    }


# ════════════════════════════════════════
# Schemas Pydantic
# ════════════════════════════════════════

class ProyectoCreate(BaseModel):
    nombre: str
    codigo: Optional[str] = None
    descripcion: Optional[str] = None


class ProyectoUpdate(BaseModel):
    nombre: Optional[str] = None
    codigo: Optional[str] = None
    descripcion: Optional[str] = None
    estado: Optional[str] = None


class ProyectoOut(BaseModel):
    id: int
    nombre: str
    codigo: Optional[str]
    descripcion: Optional[str]
    estado: str
    fecha_creacion: datetime
    fecha_actualizacion: datetime

    class Config:
        from_attributes = True


class PaqueteCreate(BaseModel):
    nodo_id: int
    nombre: Optional[str] = None
    observacion: Optional[str] = None


class PaqueteUpdate(BaseModel):
    nombre: Optional[str] = None
    observacion: Optional[str] = None


class PaqueteOut(BaseModel):
    id: int
    proyecto_id: int
    nodo_id: int
    nombre: str
    estado: str
    observacion: Optional[str] = None
    fecha_creacion: datetime
    fecha_liberacion: Optional[datetime] = None

    class Config:
        from_attributes = True


class PaqueteImpactoAPUOut(BaseModel):
    id: int
    nodo_id: int
    nombre: str
    estado: str
    rubros: int


class ImpactoAPUProyectoOut(BaseModel):
    proyecto_id: int
    apu_id: int
    total_rubros: int
    rubros_fuera_paquete: int
    paquetes: list[PaqueteImpactoAPUOut]


class NodoOut(BaseModel):
    id: int
    proyecto_id: int
    padre_id: Optional[int]
    tipo: str
    nivel: Optional[int] = None
    item: Optional[str]
    descripcion: str
    orden: int
    unidad: Optional[str]
    metrado: Optional[float]
    precio_unitario_ref: Optional[float]
    apu_id: Optional[int]
    activo_como_rubro: Optional[bool] = None
    tiene_hijos: Optional[bool] = None
    es_grupo: Optional[bool] = None
    es_rubro_operativo: Optional[bool] = None
    tipo_rubro: Optional[str]
    observaciones: Optional[str]
    individualizado: Optional[bool] = None
    estado_actualizacion: Optional[str] = None
    origen_edicion: Optional[str] = None
    requiere_revision_apu: Optional[bool] = None
    actualizacion_lote_id: Optional[int] = None
    excel_fila: Optional[int] = None
    excel_hoja: Optional[str] = None
    excel_archivo: Optional[str] = None
    fecha_actualizacion_fuente: Optional[datetime] = None
    fecha_edicion_manual: Optional[datetime] = None
    advertencia_edicion_apu: Optional[str] = None
    estado_revision_apu: Optional[str] = None
    motivos_revision_apu: list[dict[str, str]] = Field(default_factory=list)

    class Config:
        from_attributes = True


class RevisionAPUMotivoOut(BaseModel):
    codigo: str
    descripcion: str
    detalle: list[str] = Field(default_factory=list)
    aprobado: bool = False
    comentario: Optional[str] = None


class RevisionAPUHistorialItemOut(BaseModel):
    codigo_motivo: str
    descripcion_motivo: str
    aprobado: bool
    comentario: Optional[str] = None


class RevisionAPUHistorialOut(BaseModel):
    id: int
    estado: str
    items: list[RevisionAPUHistorialItemOut]


class RevisionAPUOut(BaseModel):
    nodo_id: int
    estado: str
    motivos_actuales: list[RevisionAPUMotivoOut]
    historial: list[RevisionAPUHistorialOut]


class RevisionAPUItemIn(BaseModel):
    codigo_motivo: str
    aprobado: bool
    comentario: Optional[str] = None


class RevisionAPURevisadoIn(BaseModel):
    items: list[RevisionAPUItemIn]


def _nodo_out(nodo: NodoPresupuesto, hijos_por_padre: Optional[dict[int, int]] = None) -> NodoOut:
    hijos_por_padre = hijos_por_padre or {}
    tiene_hijos = hijos_por_padre.get(nodo.id, 0) > 0
    activo = bool(nodo.activo_como_rubro) if nodo.activo_como_rubro is not None else nodo.tipo == "RUBRO"
    salida = NodoOut.model_validate(nodo)
    salida.tiene_hijos = tiene_hijos
    salida.es_grupo = tiene_hijos
    salida.es_rubro_operativo = activo and not tiene_hijos
    return salida


def _tiene_hijos(db: Session, nodo: NodoPresupuesto) -> bool:
    return (
        db.query(NodoPresupuesto.id)
        .filter(NodoPresupuesto.padre_id == nodo.id)
        .first()
        is not None
    )


def _es_rubro_operativo(db: Session, nodo: NodoPresupuesto) -> bool:
    activo = bool(nodo.activo_como_rubro) if nodo.activo_como_rubro is not None else nodo.tipo == "RUBRO"
    return activo and not _tiene_hijos(db, nodo)


def _normalizar_variante_nombre(nombre: Any) -> str:
    texto = " ".join(str(nombre or "").strip().split())
    if not texto:
        raise HTTPException(status_code=400, detail="El nombre de la variante es obligatorio")
    return texto


def _apu_base_id(apu: APU) -> int:
    return apu.apu_base_id if apu.es_variante and apu.apu_base_id else apu.id


def _apu_base(db: Session, apu: APU) -> APU:
    base_id = _apu_base_id(apu)
    base = db.query(APU).filter(APU.id == base_id).first()
    if not base:
        raise HTTPException(status_code=404, detail="APU base no encontrado")
    return base


def _siguiente_codigo_variante(db: Session, base_apu: APU) -> str:
    base_codigo = base_apu.codigo or f"APU-{base_apu.id}"
    prefijo = f"{base_codigo}-V"
    codigos = [
        row[0] for row in db.query(APU.codigo)
        .filter(APU.codigo.like(f"{prefijo}%"))
        .all()
    ]
    maximo = 0
    for codigo in codigos:
        match = re.match(rf"^{re.escape(prefijo)}(\d+)$", codigo or "")
        if match:
            maximo = max(maximo, int(match.group(1)))
    while True:
        maximo += 1
        candidato = f"{prefijo}{maximo:03d}"
        existe = db.query(APU.id).filter(APU.codigo == candidato).first()
        if not existe:
            return candidato


def _variante_out(db: Session, variante: APU) -> VarianteAPUOut:
    usos = db.query(func.count(NodoPresupuesto.id)).filter(NodoPresupuesto.apu_id == variante.id).scalar() or 0
    return VarianteAPUOut(
        id=variante.id,
        codigo=variante.codigo,
        nombre=variante.nombre,
        variante_nombre=variante.variante_nombre or "",
        apu_base_id=variante.apu_base_id or 0,
        proyecto_id=variante.proyecto_id or 0,
        usos=int(usos),
        precio_unitario=calcular_costo_apu(variante)["precio_unitario"],
    )


def _buscar_variante_existente(db: Session, proyecto_id: int, apu_base_id: int, variante_nombre: str) -> Optional[APU]:
    nombre_normalizado = _normalizar_texto(variante_nombre)
    variantes = (
        db.query(APU)
        .options(joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(
            APU.es_variante.is_(True),
            APU.proyecto_id == proyecto_id,
            APU.apu_base_id == apu_base_id,
        )
        .all()
    )
    for variante in variantes:
        if _normalizar_texto(variante.variante_nombre) == nombre_normalizado:
            return variante
    return None


def _validar_fuente_variante(db: Session, source_id: int, proyecto_id: int, apu_base_id: int) -> APU:
    source = (
        db.query(APU)
        .options(joinedload(APU.items))
        .filter(APU.id == source_id)
        .first()
    )
    if not source:
        raise HTTPException(status_code=404, detail="APU fuente no encontrado")
    if source.id == apu_base_id and not source.es_variante:
        return source
    if source.es_variante and source.apu_base_id == apu_base_id and source.proyecto_id == proyecto_id:
        return source
    raise HTTPException(status_code=400, detail="El APU fuente no corresponde al mismo APU base y proyecto")


def _copiar_items_apu(db: Session, source: APU, destino: APU) -> None:
    for item in db.query(APUItem).filter(APUItem.apu_id == source.id).all():
        db.add(
            APUItem(
                apu_id=destino.id,
                recurso_id=item.recurso_id,
                categoria=item.categoria,
                cantidad=item.cantidad,
                orden=item.orden,
                es_herramienta_menor=item.es_herramienta_menor,
            )
        )


def _codigo_recurso_proyecto(db: Session, recurso: Recurso, proyecto_id: int) -> str:
    base = f"{recurso.codigo}-P{proyecto_id}" if recurso.codigo else f"REC-{recurso.id}-P{proyecto_id}"
    candidato = base
    contador = 2
    while db.query(Recurso.id).filter(Recurso.codigo == candidato).first():
        candidato = f"{base}-{contador}"
        contador += 1
    return candidato


def _descendientes_ids(nodo_id: int, hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]]) -> set[int]:
    ids = set()
    pendientes = list(hijos_por_padre.get(nodo_id, []))
    while pendientes:
        actual = pendientes.pop()
        ids.add(actual.id)
        pendientes.extend(hijos_por_padre.get(actual.id, []))
    return ids


def _reordenar_nodos(nodos: list[NodoPresupuesto]) -> None:
    for idx, nodo in enumerate(nodos):
        nodo.orden = idx


def _aplanar_hijos(hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]]) -> list[NodoPresupuesto]:
    resultado = []

    def visitar(padre_id: Optional[int]) -> None:
        for hijo in hijos_por_padre.get(padre_id, []):
            resultado.append(hijo)
            visitar(hijo.id)

    visitar(None)
    return resultado


def _recalcular_niveles_desde_arbol(hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]]) -> None:
    def visitar(padre_id: Optional[int], nivel: int) -> None:
        for hijo in hijos_por_padre.get(padre_id, []):
            hijo.nivel = nivel
            visitar(hijo.id, nivel + 1)

    visitar(None, 0)


def _sincronizar_niveles(db: Session, proyecto_id: int) -> None:
    nodos = db.query(NodoPresupuesto).filter(NodoPresupuesto.proyecto_id == proyecto_id).all()
    if not nodos:
        return
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for n in nodos:
        hijos_por_padre.setdefault(n.padre_id, []).append(n)
    for lista in hijos_por_padre.values():
        lista.sort(key=lambda x: (x.orden or 0, x.id))
    cambios = False

    def visitar(padre_id: Optional[int], nivel: int) -> None:
        nonlocal cambios
        for hijo in hijos_por_padre.get(padre_id, []):
            if (hijo.nivel or 0) != nivel:
                hijo.nivel = nivel
                cambios = True
            visitar(hijo.id, nivel + 1)

    visitar(None, 0)
    if cambios:
        db.commit()


def _actualizar_estado_rubro_por_hijos(db: Session, nodo: Optional[NodoPresupuesto]) -> None:
    if not nodo:
        return
    if _tiene_hijos(db, nodo):
        nodo.activo_como_rubro = False
    elif nodo.activo_como_rubro is False:
        nodo.activo_como_rubro = True


class NodoUpdate(BaseModel):
    item: Optional[str] = None
    descripcion: Optional[str] = None
    unidad: Optional[str] = None
    metrado: Optional[float] = None
    precio_unitario_ref: Optional[float] = None
    observaciones: Optional[str] = None
    orden: Optional[int] = None
    estado_actualizacion: Optional[str] = None


class NodoCreate(BaseModel):
    despues_de_id: int
    descripcion: Optional[str] = None
    unidad: Optional[str] = None
    metrado: Optional[float] = None
    precio_unitario_ref: Optional[float] = None


class NodoBulkCreate(NodoCreate):
    cantidad: int = Field(ge=1, le=100)


class NodoMoverRequest(BaseModel):
    accion: str
    nodo_ids: Optional[list[int]] = None


class NodoDeleteOut(BaseModel):
    eliminados: int
    nodo_id: int
    proyecto_id: int


class NodoEstructuraSnapshot(BaseModel):
    id: int
    padre_id: Optional[int] = None
    nivel: Optional[int] = None
    orden: int
    activo_como_rubro: Optional[bool] = None


class EstructuraRestoreRequest(BaseModel):
    nodos: list[NodoEstructuraSnapshot]


class VincularAPURequest(BaseModel):
    apu_id: int


class CrearAPUDesdeRubroRequest(BaseModel):
    base_apu_id: Optional[int] = None


class CrearAPUDesdeRubroOut(BaseModel):
    nodo: NodoOut
    apu_id: int
    codigo: Optional[str]
    nombre: str
    unidad: str
    estado: str


class VarianteAPUOut(BaseModel):
    id: int
    codigo: Optional[str]
    nombre: str
    variante_nombre: str
    apu_base_id: int
    proyecto_id: int
    usos: int
    precio_unitario: Optional[float] = None


class CrearVarianteAPURequest(BaseModel):
    variante_nombre: str
    copiar_desde_apu_id: Optional[int] = None


class CambiarVarianteAPURequest(BaseModel):
    variante_apu_id: Optional[int] = None


class VarianteAPUAsignadaOut(BaseModel):
    nodo: NodoOut
    variante: Optional[VarianteAPUOut] = None
    apu_id: int
    created: bool = False


# ════════════════════════════════════════
# CRUD Proyectos
# ════════════════════════════════════════

class AislarAPUNoLiberadosOut(BaseModel):
    variante: VarianteAPUOut
    apu_id: int
    rubros_reasignados: int
    rubros_liberados_preservados: int
    created: bool = True


class RecursoProyectoUpdate(BaseModel):
    descripcion: str
    unidad: str
    precio_unitario: float
    fuente_precio: Optional[str] = None
    observacion: Optional[str] = None
    estado_validacion: str = "pendiente"
    nota_validacion: Optional[str] = None


class RecursoProyectoOut(BaseModel):
    recurso: RecursoOut
    recurso_id: int
    apu_id: int
    created: bool = True


@router.get("/proyectos/", response_model=list[ProyectoOut])
def listar_proyectos(background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    if not backup.existe_backup_diario_hoy():
        background_tasks.add_task(backup.crear_backup, "auto_diario", "diario")
    return db.query(Proyecto).order_by(Proyecto.fecha_creacion.desc()).all()


@router.post("/proyectos/", response_model=ProyectoOut, status_code=201)
def crear_proyecto(data: ProyectoCreate, db: Session = Depends(get_db)):
    # Verificar código único si se proporcionó
    if data.codigo:
        existe = db.query(Proyecto).filter(Proyecto.codigo == data.codigo).first()
        if existe:
            raise HTTPException(
                status_code=400, detail=f"Ya existe un proyecto con código '{data.codigo}'"
            )
    proyecto = Proyecto(
        nombre=data.nombre,
        codigo=data.codigo,
        descripcion=data.descripcion,
    )
    db.add(proyecto)
    db.commit()
    db.refresh(proyecto)
    return proyecto


@router.get("/proyectos/{proyecto_id}", response_model=ProyectoOut)
def obtener_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    return proyecto


@router.put("/proyectos/{proyecto_id}", response_model=ProyectoOut)
def actualizar_proyecto(
    proyecto_id: int, data: ProyectoUpdate, db: Session = Depends(get_db)
):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    if data.nombre is not None:
        proyecto.nombre = data.nombre
    if data.codigo is not None:
        # Verificar que el nuevo código no lo usa otro proyecto
        otro = (
            db.query(Proyecto)
            .filter(Proyecto.codigo == data.codigo, Proyecto.id != proyecto_id)
            .first()
        )
        if otro:
            raise HTTPException(
                status_code=400, detail=f"Código '{data.codigo}' ya está en uso"
            )
        proyecto.codigo = data.codigo
    if data.descripcion is not None:
        proyecto.descripcion = data.descripcion
    if data.estado is not None:
        if data.estado not in ("activo", "archivado"):
            raise HTTPException(
                status_code=400, detail="Estado debe ser 'activo' o 'archivado'"
            )
        proyecto.estado = data.estado

    proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()
    db.refresh(proyecto)
    return proyecto


@router.delete("/proyectos/{proyecto_id}", status_code=204)
def eliminar_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    db.delete(proyecto)
    db.commit()


# ════════════════════════════════════════
# Nodos — árbol completo
# ════════════════════════════════════════

# Paquetes de presupuesto

def _validar_nodo_paquete(db: Session, proyecto_id: int, nodo_id: int) -> NodoPresupuesto:
    proyecto = db.query(Proyecto.id).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    nodo = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.id == nodo_id, NodoPresupuesto.proyecto_id == proyecto_id)
        .first()
    )
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo seleccionado no encontrado en el proyecto")
    return nodo


@router.get("/proyectos/{proyecto_id}/paquetes", response_model=list[PaqueteOut])
def listar_paquetes(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto.id).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    return (
        db.query(PaquetePresupuesto)
        .filter(PaquetePresupuesto.proyecto_id == proyecto_id)
        .order_by(PaquetePresupuesto.fecha_creacion.desc(), PaquetePresupuesto.id.desc())
        .all()
    )


@router.post("/proyectos/{proyecto_id}/paquetes", response_model=PaqueteOut, status_code=201)
def crear_paquete(proyecto_id: int, data: PaqueteCreate, db: Session = Depends(get_db)):
    nodo = _validar_nodo_paquete(db, proyecto_id, data.nodo_id)
    existente = (
        db.query(PaquetePresupuesto)
        .filter(PaquetePresupuesto.proyecto_id == proyecto_id, PaquetePresupuesto.nodo_id == data.nodo_id)
        .first()
    )
    if existente:
        raise HTTPException(status_code=400, detail="La rama seleccionada ya tiene un paquete definido")

    nombre = " ".join((data.nombre or nodo.descripcion or "").split())
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre del paquete es obligatorio")

    paquete = PaquetePresupuesto(
        proyecto_id=proyecto_id,
        nodo_id=data.nodo_id,
        nombre=nombre[:240],
        observacion=data.observacion,
    )
    db.add(paquete)
    db.commit()
    db.refresh(paquete)
    return paquete


@router.patch("/paquetes/{paquete_id}", response_model=PaqueteOut)
def actualizar_paquete(paquete_id: int, data: PaqueteUpdate, db: Session = Depends(get_db)):
    paquete = db.query(PaquetePresupuesto).filter(PaquetePresupuesto.id == paquete_id).first()
    if not paquete:
        raise HTTPException(status_code=404, detail="Paquete no encontrado")
    if data.nombre is not None:
        nombre = " ".join(data.nombre.split())
        if not nombre:
            raise HTTPException(status_code=400, detail="El nombre del paquete es obligatorio")
        paquete.nombre = nombre[:240]
    if data.observacion is not None:
        paquete.observacion = data.observacion
    db.commit()
    db.refresh(paquete)
    return paquete


@router.patch("/paquetes/{paquete_id}/liberar", response_model=PaqueteOut)
def liberar_paquete(paquete_id: int, db: Session = Depends(get_db)):
    paquete = db.query(PaquetePresupuesto).filter(PaquetePresupuesto.id == paquete_id).first()
    if not paquete:
        raise HTTPException(status_code=404, detail="Paquete no encontrado")
    paquete.estado = "liberado"
    paquete.fecha_liberacion = datetime.utcnow()
    db.commit()
    db.refresh(paquete)
    return paquete


@router.patch("/paquetes/{paquete_id}/reabrir", response_model=PaqueteOut)
def reabrir_paquete(paquete_id: int, db: Session = Depends(get_db)):
    paquete = db.query(PaquetePresupuesto).filter(PaquetePresupuesto.id == paquete_id).first()
    if not paquete:
        raise HTTPException(status_code=404, detail="Paquete no encontrado")
    paquete.estado = "activo"
    paquete.fecha_liberacion = None
    db.commit()
    db.refresh(paquete)
    return paquete


@router.get("/proyectos/{proyecto_id}/apus/{apu_id}/impacto-paquetes", response_model=ImpactoAPUProyectoOut)
def impacto_apu_paquetes(proyecto_id: int, apu_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto.id).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    apu = db.query(APU.id).filter(APU.id == apu_id).first()
    if not apu:
        raise HTTPException(status_code=404, detail="APU no encontrado")

    nodos = (
        db.query(NodoPresupuesto.id, NodoPresupuesto.padre_id, NodoPresupuesto.apu_id)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .all()
    )
    padre_por_id = {nodo.id: nodo.padre_id for nodo in nodos}
    usos = [nodo for nodo in nodos if nodo.apu_id == apu_id]
    paquetes = (
        db.query(PaquetePresupuesto)
        .filter(PaquetePresupuesto.proyecto_id == proyecto_id)
        .all()
    )
    paquete_por_nodo = {paquete.nodo_id: paquete for paquete in paquetes}
    conteo_por_paquete: dict[int, int] = {}
    fuera_paquete = 0

    for uso in usos:
        cursor = uso.id
        paquete_encontrado: Optional[PaquetePresupuesto] = None
        while cursor is not None:
            paquete_encontrado = paquete_por_nodo.get(cursor)
            if paquete_encontrado:
                break
            cursor = padre_por_id.get(cursor)
        if paquete_encontrado:
            conteo_por_paquete[paquete_encontrado.id] = conteo_por_paquete.get(paquete_encontrado.id, 0) + 1
        else:
            fuera_paquete += 1

    paquetes_impactados = [
        PaqueteImpactoAPUOut(
            id=paquete.id,
            nodo_id=paquete.nodo_id,
            nombre=paquete.nombre,
            estado=paquete.estado,
            rubros=conteo,
        )
        for paquete in paquetes
        if (conteo := conteo_por_paquete.get(paquete.id, 0)) > 0
    ]
    paquetes_impactados.sort(key=lambda item: (item.estado != "liberado", item.nombre.lower()))

    return ImpactoAPUProyectoOut(
        proyecto_id=proyecto_id,
        apu_id=apu_id,
        total_rubros=len(usos),
        rubros_fuera_paquete=fuera_paquete,
        paquetes=paquetes_impactados,
    )


@router.get("/proyectos/{proyecto_id}/nodos", response_model=list[NodoOut])
def listar_nodos(proyecto_id: int, db: Session = Depends(get_db)):
    """
    Devuelve todos los nodos del proyecto ordenados por 'orden'.
    El frontend construye el árbol usando padre_id.
    """
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    _sincronizar_niveles(db, proyecto_id)

    nodos = (
        db.query(NodoPresupuesto)
        .options(joinedload(NodoPresupuesto.apu).joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .order_by(NodoPresupuesto.orden)
        .all()
    )
    hijos_por_padre = dict(
        db.query(NodoPresupuesto.padre_id, func.count(NodoPresupuesto.id))
        .filter(NodoPresupuesto.proyecto_id == proyecto_id, NodoPresupuesto.padre_id.isnot(None))
        .group_by(NodoPresupuesto.padre_id)
        .all()
    )
    costos_por_apu: dict[int, dict] = {}
    salida = []
    for nodo in nodos:
        item = _nodo_out(nodo, hijos_por_padre)
        if item.es_rubro_operativo and nodo.apu:
            if nodo.apu.id not in costos_por_apu:
                costos_por_apu[nodo.apu.id] = _costo_apu_con_control(nodo.apu) or {}
            costo_apu = costos_por_apu[nodo.apu.id]
            item.motivos_revision_apu = _motivos_revision_apu(nodo, costo_apu)
            item.estado_revision_apu = _estado_revision_apu(nodo, costo_apu, db)
        salida.append(item)
    return salida


@router.get("/proyectos/{proyecto_id}/exportar-operativo.xlsx")
def exportar_presupuesto_operativo(
    proyecto_id: int,
    db: Session = Depends(get_db),
    root_nodo_id: Optional[int] = None,
    vista_exportacion: Optional[str] = None,
):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    _sincronizar_niveles(db, proyecto_id)

    nodos = (
        db.query(NodoPresupuesto)
        .options(joinedload(NodoPresupuesto.apu).joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    if root_nodo_id is not None:
        if not any(n.id == root_nodo_id for n in nodos):
            raise HTTPException(status_code=404, detail="Nodo seleccionado no encontrado en el proyecto")
        hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
        for nodo in nodos:
            hijos_por_padre.setdefault(nodo.padre_id, []).append(nodo)
        ids_exportar = {root_nodo_id, *_descendientes_ids(root_nodo_id, hijos_por_padre)}
        nodos = [n for n in nodos if n.id in ids_exportar]
    nodos = _filtrar_nodos_por_vista_exportacion(nodos, vista_exportacion, db)
    archivo = _crear_workbook_presupuesto_operativo(proyecto, nodos, db)
    filename = _nombre_archivo_excel(proyecto)
    return StreamingResponse(
        archivo,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/nodos/{nodo_id}", response_model=NodoOut)
def editar_nodo(nodo_id: int, data: NodoUpdate, db: Session = Depends(get_db)):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Por ahora solo se editan nodos operativos tipo rubro")
    if data.estado_actualizacion is not None and data.estado_actualizacion not in ("activo", "obsoleto"):
        raise HTTPException(status_code=400, detail="Estado debe ser 'activo' u 'obsoleto'")

    cambia_nombre = data.descripcion is not None and _distinto_texto(nodo.descripcion, data.descripcion)
    cambia_unidad = data.unidad is not None and _distinto_texto(nodo.unidad, data.unidad)
    advertencia = None

    campos = data.model_fields_set

    if "item" in campos:
        raise HTTPException(status_code=400, detail="El item/codigo es automatico y no se edita manualmente")
    if "descripcion" in campos:
        texto = _texto(data.descripcion)
        if not texto:
            raise HTTPException(status_code=400, detail="La descripcion no puede quedar vacia")
        nodo.descripcion = texto
    if "unidad" in campos:
        nodo.unidad = _texto(data.unidad)
    if "metrado" in campos:
        nodo.metrado = float(data.metrado) if data.metrado is not None else None
    if "precio_unitario_ref" in campos:
        nodo.precio_unitario_ref = float(data.precio_unitario_ref) if data.precio_unitario_ref is not None else None
    if "observaciones" in campos:
        nodo.observaciones = _texto(data.observaciones)
    if "orden" in campos:
        nodo.orden = int(data.orden) if data.orden is not None else nodo.orden
    if "estado_actualizacion" in campos:
        nodo.estado_actualizacion = data.estado_actualizacion

    if nodo.apu_id and (cambia_nombre or cambia_unidad):
        nodo.requiere_revision_apu = True
        advertencia = "El rubro tiene APU vinculado; revisa compatibilidad de descripcion/unidad."

    nodo.origen_edicion = "manual"
    nodo.fecha_edicion_manual = datetime.utcnow()
    proyecto = db.query(Proyecto).filter(Proyecto.id == nodo.proyecto_id).first()
    if proyecto:
        proyecto.fecha_actualizacion = datetime.utcnow()

    db.commit()
    db.refresh(nodo)
    salida = _nodo_out(nodo)
    salida.advertencia_edicion_apu = advertencia
    return salida


def _crear_rubros_debajo(
    db: Session,
    proyecto: Proyecto,
    referencia: NodoPresupuesto,
    data: NodoCreate,
    cantidad: int,
) -> list[NodoPresupuesto]:
    orden_referencia = referencia.orden or 0
    (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto.id, NodoPresupuesto.orden > orden_referencia)
        .update({NodoPresupuesto.orden: NodoPresupuesto.orden + cantidad}, synchronize_session=False)
    )

    nuevos = []
    fecha = datetime.utcnow()
    for indice in range(cantidad):
        nuevo = NodoPresupuesto(
            proyecto_id=proyecto.id,
            padre_id=referencia.padre_id,
            tipo="RUBRO",
            nivel=referencia.nivel if referencia.nivel is not None else 6,
            item=None,
            descripcion=_texto(data.descripcion) or "(nuevo rubro)",
            orden=orden_referencia + 1 + indice,
            unidad=_texto(data.unidad),
            metrado=float(data.metrado) if data.metrado is not None else None,
            precio_unitario_ref=float(data.precio_unitario_ref) if data.precio_unitario_ref is not None else None,
            activo_como_rubro=True,
            tipo_rubro="PENDIENTE",
            observaciones=None,
            estado_actualizacion="activo",
            origen_edicion="manual",
            requiere_revision_apu=False,
            fecha_edicion_manual=fecha,
        )
        db.add(nuevo)
        nuevos.append(nuevo)

    db.flush()
    _renumerar_rubros_activos_hermanos(db, proyecto.id, referencia.padre_id)
    db.flush()
    proyecto.fecha_actualizacion = datetime.utcnow()
    return nuevos


@router.post("/proyectos/{proyecto_id}/nodos", response_model=NodoOut, status_code=201)
def crear_rubro_debajo(proyecto_id: int, data: NodoCreate, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    referencia = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.id == data.despues_de_id, NodoPresupuesto.proyecto_id == proyecto_id)
        .first()
    )
    if not referencia:
        raise HTTPException(status_code=404, detail="Nodo de referencia no encontrado")
    if not _es_rubro_operativo(db, referencia):
        raise HTTPException(status_code=400, detail="La fila nueva debe agregarse debajo de un rubro operativo")

    nuevo = _crear_rubros_debajo(db, proyecto, referencia, data, 1)[0]
    db.commit()
    db.refresh(nuevo)
    return _nodo_out(nuevo)


@router.post("/proyectos/{proyecto_id}/nodos/bulk", response_model=list[NodoOut], status_code=201)
def crear_rubros_debajo(proyecto_id: int, data: NodoBulkCreate, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    referencia = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.id == data.despues_de_id, NodoPresupuesto.proyecto_id == proyecto_id)
        .first()
    )
    if not referencia:
        raise HTTPException(status_code=404, detail="Nodo de referencia no encontrado")
    if not _es_rubro_operativo(db, referencia):
        raise HTTPException(status_code=400, detail="Las filas nuevas deben agregarse debajo de un rubro operativo")

    nuevos = _crear_rubros_debajo(db, proyecto, referencia, data, data.cantidad)
    db.commit()
    for nuevo in nuevos:
        db.refresh(nuevo)
    return [_nodo_out(nuevo) for nuevo in nuevos]


@router.patch("/nodos/{nodo_id}/marcar-obsoleto", response_model=NodoOut)
def marcar_obsoleto(nodo_id: int, db: Session = Depends(get_db)):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se pueden marcar obsoletos nodos operativos tipo rubro")

    nodo.estado_actualizacion = "obsoleto"
    nodo.origen_edicion = "manual"
    nodo.fecha_edicion_manual = datetime.utcnow()
    db.flush()
    _renumerar_rubros_activos_hermanos(db, nodo.proyecto_id, nodo.padre_id)
    proyecto = db.query(Proyecto).filter(Proyecto.id == nodo.proyecto_id).first()
    if proyecto:
        proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)


# ════════════════════════════════════════
# Importación desde Excel
# ════════════════════════════════════════

@router.delete("/nodos/{nodo_id}/bloque", response_model=NodoDeleteOut)
def eliminar_bloque_nodo(nodo_id: int, db: Session = Depends(get_db)):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")

    proyecto_id = nodo.proyecto_id
    padre_id = nodo.padre_id
    nodos = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    por_id = {n.id: n for n in nodos}
    for item in nodos:
        hijos_por_padre.setdefault(item.padre_id, []).append(item)

    bloque_ids = {nodo_id, *_descendientes_ids(nodo_id, hijos_por_padre)}
    backup.crear_backup(f"eliminar_bloque_proyecto_{proyecto_id}_{nodo_id}", tipo="critico")

    for bloque_id in sorted(bloque_ids, key=lambda item_id: por_id[item_id].nivel or 0, reverse=True):
        db.delete(por_id[bloque_id])

    db.flush()
    _renumerar_rubros_activos_hermanos(db, proyecto_id, padre_id)
    _actualizar_estado_rubro_por_hijos(db, por_id.get(padre_id) if padre_id else None)

    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if proyecto:
        proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()

    return NodoDeleteOut(eliminados=len(bloque_ids), nodo_id=nodo_id, proyecto_id=proyecto_id)


@router.patch("/nodos/{nodo_id}/mover-estructura", response_model=list[NodoOut])
def mover_estructura(nodo_id: int, data: NodoMoverRequest, db: Session = Depends(get_db)):
    accion = (data.accion or "").strip().lower()
    if accion not in ("subir", "bajar", "sangrar", "agrupar_abajo", "quitar_sangria", "quitar_sangria_con_bloque"):
        raise HTTPException(status_code=400, detail="Accion invalida")

    nodo_base = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo_base:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")

    nodos = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == nodo_base.proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    por_id = {n.id: n for n in nodos}
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for n in nodos:
        hijos_por_padre.setdefault(n.padre_id, []).append(n)

    ids_solicitados = data.nodo_ids or [nodo_id]
    ids_unicos = []
    for seleccion_id in ids_solicitados:
        if seleccion_id not in ids_unicos:
            ids_unicos.append(seleccion_id)
    if not ids_unicos:
        raise HTTPException(status_code=400, detail="Selecciona al menos una fila")

    seleccion = []
    for seleccion_id in ids_unicos:
        nodo = por_id.get(seleccion_id)
        if not nodo or nodo.proyecto_id != nodo_base.proyecto_id:
            raise HTTPException(status_code=404, detail="Alguna fila seleccionada no existe en este proyecto")
        seleccion.append(nodo)

    seleccion_ids = {n.id for n in seleccion}
    descendientes_por_nodo = {n.id: _descendientes_ids(n.id, hijos_por_padre) for n in seleccion}
    for n in seleccion:
        if descendientes_por_nodo[n.id] & seleccion_ids:
            raise HTTPException(
                status_code=400,
                detail="La seleccion mezcla un grupo con sus hijos. Selecciona solo el grupo o solo sus filas hijas.",
            )

    padres = {n.padre_id for n in seleccion}
    if len(padres) != 1:
        raise HTTPException(status_code=400, detail="Selecciona filas contiguas del mismo nivel")

    padre_anterior_id = seleccion[0].padre_id
    hermanos = hijos_por_padre.get(padre_anterior_id, [])
    indices = sorted(next((i for i, h in enumerate(hermanos) if h.id == n.id), -1) for n in seleccion)
    if any(i < 0 for i in indices) or indices != list(range(indices[0], indices[-1] + 1)):
        raise HTTPException(status_code=400, detail="Selecciona filas contiguas del mismo nivel")

    seleccion_ordenada = [hermanos[i] for i in indices]
    bloque_ids = set(seleccion_ids)
    for ids in descendientes_por_nodo.values():
        bloque_ids.update(ids)
    padres_para_actualizar: set[Optional[int]] = {padre_anterior_id}

    def extraer_bloque_de_hermanos() -> list[NodoPresupuesto]:
        return [n for n in hermanos if n.id in seleccion_ids]

    if accion in ("subir", "bajar"):
        inicio = indices[0]
        fin = indices[-1]
        if (accion == "subir" and inicio == 0) or (accion == "bajar" and fin >= len(hermanos) - 1):
            raise HTTPException(status_code=400, detail="La fila no tiene hermano disponible en esa direccion dentro de su grupo")
        bloque = extraer_bloque_de_hermanos()
        if accion == "subir":
            hermano = hermanos[inicio - 1]
            hijos_por_padre[padre_anterior_id] = [
                *hermanos[:inicio - 1],
                *bloque,
                hermano,
                *hermanos[fin + 1:],
            ]
        else:
            hermano = hermanos[fin + 1]
            hijos_por_padre[padre_anterior_id] = [
                *hermanos[:inicio],
                hermano,
                *bloque,
                *hermanos[fin + 2:],
            ]

    elif accion == "sangrar":
        inicio = indices[0]
        if inicio <= 0:
            raise HTTPException(status_code=400, detail="No hay nodo anterior para usar como grupo")
        nuevo_padre = hermanos[inicio - 1]
        if any(nuevo_padre.id in descendientes_por_nodo[n.id] for n in seleccion_ordenada):
            raise HTTPException(status_code=400, detail="No se puede mover un nodo dentro de sus propios hijos")
        delta = (nuevo_padre.nivel or 0) + 1 - (seleccion_ordenada[0].nivel or 0)
        if any(((por_id[i].nivel or 0) + delta) > 7 for i in bloque_ids):
            raise HTTPException(status_code=400, detail="Nivel maximo de jerarquia alcanzado (8 niveles)")
        hijos_por_padre[padre_anterior_id] = [n for n in hermanos if n.id not in seleccion_ids]
        hijos_destino = hijos_por_padre.setdefault(nuevo_padre.id, [])
        for n in seleccion_ordenada:
            n.padre_id = nuevo_padre.id
            hijos_destino.append(n)
        padres_para_actualizar.add(nuevo_padre.id)
        nuevo_padre.activo_como_rubro = False

    elif accion == "agrupar_abajo":
        if len(seleccion_ordenada) < 2:
            raise HTTPException(status_code=400, detail="Selecciona la fila y al menos una fila de abajo para agrupar")
        nuevo_padre = seleccion_ordenada[0]
        hijos_a_mover = seleccion_ordenada[1:]
        hijos_ids = {n.id for n in hijos_a_mover}
        delta = (nuevo_padre.nivel or 0) + 1 - (hijos_a_mover[0].nivel or 0)
        hijos_bloque_ids = set(hijos_ids)
        for n in hijos_a_mover:
            hijos_bloque_ids.update(descendientes_por_nodo[n.id])
        if any(((por_id[i].nivel or 0) + delta) > 7 for i in hijos_bloque_ids):
            raise HTTPException(status_code=400, detail="Nivel maximo de jerarquia alcanzado (8 niveles)")
        hijos_por_padre[padre_anterior_id] = [n for n in hermanos if n.id not in hijos_ids]
        hijos_destino = hijos_por_padre.setdefault(nuevo_padre.id, [])
        for n in hijos_a_mover:
            n.padre_id = nuevo_padre.id
            hijos_destino.append(n)
        padres_para_actualizar.add(nuevo_padre.id)
        nuevo_padre.activo_como_rubro = False

    elif accion in ("quitar_sangria", "quitar_sangria_con_bloque"):
        padre = por_id.get(padre_anterior_id)
        if not padre:
            raise HTTPException(status_code=400, detail="La seleccion ya esta en el nivel superior")
        if accion == "quitar_sangria_con_bloque" and len(seleccion_ordenada) < 2:
            raise HTTPException(status_code=400, detail="Selecciona la fila y al menos una fila de abajo para agrupar")
        hermanos_destino = hijos_por_padre.setdefault(padre.padre_id, [])
        idx_padre = next((i for i, h in enumerate(hermanos_destino) if h.id == padre.id), len(hermanos_destino) - 1)
        if accion == "quitar_sangria":
            hijos_por_padre[padre_anterior_id] = [n for n in hermanos if n.id not in seleccion_ids]
            for offset, n in enumerate(seleccion_ordenada, start=1):
                n.padre_id = padre.padre_id
                hermanos_destino.insert(idx_padre + offset, n)
            padres_para_actualizar.update({padre.id, padre.padre_id})
        else:
            nuevo_padre = seleccion_ordenada[0]
            hijos_a_mover = seleccion_ordenada[1:]
            hijos_ids = {n.id for n in hijos_a_mover}
            delta = (nuevo_padre.nivel or 0) + 1 - (hijos_a_mover[0].nivel or 0)
            hijos_bloque_ids = set(hijos_ids)
            for n in hijos_a_mover:
                hijos_bloque_ids.update(descendientes_por_nodo[n.id])
            if any(((por_id[i].nivel or 0) + delta) > 7 for i in hijos_bloque_ids):
                raise HTTPException(status_code=400, detail="Nivel maximo de jerarquia alcanzado (8 niveles)")
            hijos_por_padre[padre_anterior_id] = [n for n in hermanos if n.id not in seleccion_ids]
            nuevo_padre.padre_id = padre.padre_id
            hermanos_destino.insert(idx_padre + 1, nuevo_padre)
            hijos_destino = hijos_por_padre.setdefault(nuevo_padre.id, [])
            for n in hijos_a_mover:
                n.padre_id = nuevo_padre.id
                hijos_destino.append(n)
            padres_para_actualizar.update({padre.id, padre.padre_id, nuevo_padre.id})
            nuevo_padre.activo_como_rubro = False

    _recalcular_niveles_desde_arbol(hijos_por_padre)
    nodos = _aplanar_hijos(hijos_por_padre)
    _reordenar_nodos(nodos)
    db.flush()
    for padre_id in padres_para_actualizar:
        _actualizar_estado_rubro_por_hijos(db, por_id.get(padre_id) if padre_id else None)

    proyecto = db.query(Proyecto).filter(Proyecto.id == nodo_base.proyecto_id).first()
    if proyecto:
        proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()

    hijos_counts = dict(
        db.query(NodoPresupuesto.padre_id, func.count(NodoPresupuesto.id))
        .filter(NodoPresupuesto.proyecto_id == nodo_base.proyecto_id, NodoPresupuesto.padre_id.isnot(None))
        .group_by(NodoPresupuesto.padre_id)
        .all()
    )
    actualizados = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == nodo_base.proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    return [_nodo_out(n, hijos_counts) for n in actualizados]


@router.patch("/proyectos/{proyecto_id}/nodos/estructura", response_model=list[NodoOut])
def restaurar_estructura(proyecto_id: int, data: EstructuraRestoreRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not data.nodos:
        raise HTTPException(status_code=400, detail="No hay estructura para restaurar")

    nodos = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .all()
    )
    por_id = {n.id: n for n in nodos}
    snapshot_ids = {n.id for n in data.nodos}
    if snapshot_ids != set(por_id):
        raise HTTPException(status_code=400, detail="La estructura guardada ya no coincide con el presupuesto actual")

    for snap in data.nodos:
        if snap.padre_id is not None and snap.padre_id not in por_id:
            raise HTTPException(status_code=400, detail="La estructura guardada contiene un padre inexistente")
        nodo = por_id[snap.id]
        nodo.padre_id = snap.padre_id
        nodo.nivel = snap.nivel
        nodo.orden = snap.orden
        nodo.activo_como_rubro = snap.activo_como_rubro

    proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()

    hijos_counts = dict(
        db.query(NodoPresupuesto.padre_id, func.count(NodoPresupuesto.id))
        .filter(NodoPresupuesto.proyecto_id == proyecto_id, NodoPresupuesto.padre_id.isnot(None))
        .group_by(NodoPresupuesto.padre_id)
        .all()
    )
    actualizados = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
        .all()
    )
    return [_nodo_out(n, hijos_counts) for n in actualizados]


@router.post("/proyectos/{proyecto_id}/importar", status_code=201)
def importar_excel(
    proyecto_id: int,
    archivo: UploadFile = File(...),
    hoja: str = Form(default="PPTO META"),
    db: Session = Depends(get_db),
):
    """
    Importa un presupuesto desde Excel.

    Parámetros:
      - archivo: el .xlsx a importar
      - hoja: nombre de la hoja a leer (default: "PPTO META")

    La hoja debe tener columnas en este orden desde la fila 6:
      Tipo | Item | Descripción | Und. | Metrado | P.U. | P.Total

    Tipos válidos: FASE, CATEGORIA, SUBCATEGORIA, CAPITULO, SUBCAPITULO, GRUPO, RUBRO

    Comportamiento con rubros SIN_APU:
      - precio_unitario_ref = 0
      - tipo_rubro = PENDIENTE
      - observaciones = "SIN_APU"
    """
    # 1. Verificar que el proyecto existe
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    # 2. Verificar que el proyecto no tenga nodos ya (evitar doble importación)
    nodos_existentes = (
        db.query(NodoPresupuesto)
        .filter(NodoPresupuesto.proyecto_id == proyecto_id)
        .count()
    )
    if nodos_existentes > 0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"El proyecto ya tiene {nodos_existentes} nodos importados. "
                "Elimínalos antes de reimportar."
            ),
        )

    backup.crear_backup(f"importar_excel_proyecto_{proyecto_id}", tipo="critico")

    # 3. Leer el archivo Excel
    try:
        import openpyxl

        contenido = archivo.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"No se pudo leer el archivo Excel: {e}"
        )

    if hoja not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"Hoja '{hoja}' no encontrada. Hojas disponibles: {wb.sheetnames}",
        )

    ws = wb[hoja]

    # 4. Leer filas (datos reales empiezan en fila 6)
    filas_datos = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        tipo = row[0]
        if tipo is None:
            continue
        tipo = str(tipo).strip().upper()
        if tipo not in ORDEN_JERARQUIA:
            continue  # Ignorar filas con tipo desconocido
        filas_datos.append(row)

    if not filas_datos:
        raise HTTPException(
            status_code=400,
            detail="No se encontraron filas válidas en la hoja especificada.",
        )

    # 5. Construir nodos y asignar padre_id usando stack
    #    El stack guarda (tipo, nodo_db) del ancestro más reciente de cada nivel
    nodos_creados = []
    stack: list[tuple[str, NodoPresupuesto]] = []

    for orden_global, row in enumerate(filas_datos):
        tipo = str(row[0]).strip().upper()
        item = str(row[1]).strip() if row[1] else None
        descripcion = str(row[2]).strip() if row[2] else "(sin descripción)"
        unidad = str(row[3]).strip() if row[3] else None
        metrado_raw = row[4]
        pu_raw = row[5]

        # ── Determinar padre ────────────────────────────────────
        nivel_actual = ORDEN_JERARQUIA.index(tipo)

        # Sacar del stack todos los nodos con nivel >= al actual
        while stack and ORDEN_JERARQUIA.index(stack[-1][0]) >= nivel_actual:
            stack.pop()

        padre_nodo = stack[-1][1] if stack else None

        # ── Campos específicos de RUBROs ─────────────────────────
        if tipo == "RUBRO":
            # Metrado
            if isinstance(metrado_raw, (int, float)):
                metrado = float(metrado_raw)
            else:
                metrado = None

            # Precio unitario y tipo_rubro
            if pu_raw == "SIN_APU" or (
                isinstance(pu_raw, str) and pu_raw.strip().upper() == "SIN_APU"
            ):
                precio_ref = 0.0
                tipo_rubro = "PENDIENTE"
                observaciones = "SIN_APU"
            elif isinstance(pu_raw, (int, float)):
                precio_ref = float(pu_raw)
                tipo_rubro = "PENDIENTE"
                observaciones = None
            else:
                # Valor inesperado (fórmula de texto, etc.)
                precio_ref = 0.0
                tipo_rubro = "PENDIENTE"
                observaciones = str(pu_raw) if pu_raw else None
        else:
            metrado = None
            precio_ref = None
            tipo_rubro = None
            observaciones = None
            unidad = None

        # ── Crear nodo ───────────────────────────────────────────
        nodo = NodoPresupuesto(
            proyecto_id=proyecto_id,
            padre_id=padre_nodo.id if padre_nodo else None,
            tipo=tipo,
            nivel=nivel_actual,
            item=item,
            descripcion=descripcion,
            orden=orden_global,
            unidad=unidad,
            metrado=metrado,
            precio_unitario_ref=precio_ref,
            activo_como_rubro=(tipo == "RUBRO"),
            tipo_rubro=tipo_rubro,
            observaciones=observaciones,
        )
        db.add(nodo)
        db.flush()  # Obtener el id sin hacer commit, necesario para padre_id de los hijos

        nodos_creados.append(nodo)

        # Agregar al stack solo si no es RUBRO (los rubros no tienen hijos)
        if tipo != "RUBRO":
            stack.append((tipo, nodo))

    # 6. Confirmar todos los nodos en una sola transacción
    db.commit()

    # 7. Estadísticas para respuesta
    conteo_por_tipo = {}
    for n in nodos_creados:
        conteo_por_tipo[n.tipo] = conteo_por_tipo.get(n.tipo, 0) + 1

    return {
        "mensaje": f"Importación completada: {len(nodos_creados)} nodos creados",
        "proyecto_id": proyecto_id,
        "hoja_importada": hoja,
        "total_nodos": len(nodos_creados),
        "por_tipo": conteo_por_tipo,
    }


# ════════════════════════════════════════
# Vincular / Desvincular APU
# ════════════════════════════════════════

# ════════════════════════════════════════
# Actualizacion desde Excel
# ════════════════════════════════════════

@router.post("/proyectos/{proyecto_id}/actualizaciones/preview")
def preview_actualizacion_excel(
    proyecto_id: int,
    archivo: UploadFile = File(...),
    hoja: str = Form(default="PPTO 260615"),
    db: Session = Depends(get_db),
):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    contenido = archivo.file.read()
    preview = _construir_preview_actualizacion(
        db=db,
        proyecto_id=proyecto_id,
        contenido=contenido,
        archivo_nombre=archivo.filename,
        hoja=hoja,
    )
    preview.pop("_filas_excel", None)
    return preview


@router.post("/proyectos/{proyecto_id}/actualizaciones/apply")
def aplicar_actualizacion_excel(
    proyecto_id: int,
    archivo: UploadFile = File(...),
    hoja: str = Form(default="PPTO 260615"),
    db: Session = Depends(get_db),
):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    contenido = archivo.file.read()
    preview = _construir_preview_actualizacion(
        db=db,
        proyecto_id=proyecto_id,
        contenido=contenido,
        archivo_nombre=archivo.filename,
        hoja=hoja,
    )
    filas_excel = preview["_filas_excel"]
    resumen = preview["resumen"]
    safe_items = {
        r["item"]
        for r in preview["paquetes"]["automatico_seguro"]["rubros_existentes"]
    }
    obsolete_items = {
        r["item"]
        for r in preview["paquetes"]["obsoletos"]
    }
    nuevos_items = {
        rubro["item"]
        for paquete in preview["paquetes"]["automatico_seguro"]["nuevos_por_ancla"]
        for rubro in paquete["rubros"]
    }
    items_a_crear = set(nuevos_items)
    for fila in filas_excel:
        if fila["tipo"] == "RUBRO" and fila["item"] in nuevos_items:
            for ancestro in fila["path"]:
                if ancestro["item"]:
                    items_a_crear.add(ancestro["item"])

    try:
        lote = ActualizacionPresupuestoLote(
            proyecto_id=proyecto_id,
            archivo=archivo.filename,
            hoja=hoja,
            estado="aplicado",
            total_nodos_excel=resumen["total_nodos_excel"],
            total_rubros_excel=resumen["total_rubros_excel"],
            total_nodos_antes=resumen["total_nodos_actual"],
            total_rubros_antes=resumen["total_rubros_actual"],
            total_excepciones=resumen["rubros_revision"],
            resumen_json=json.dumps(resumen, ensure_ascii=False),
            fecha_creacion=datetime.utcnow(),
        )
        db.add(lote)
        db.flush()

        nodos = (
            db.query(NodoPresupuesto)
            .filter(NodoPresupuesto.proyecto_id == proyecto_id)
            .order_by(NodoPresupuesto.orden, NodoPresupuesto.id)
            .all()
        )
        nodos_por_item = {n.item: n for n in nodos if n.item}
        max_orden = max([n.orden or 0 for n in nodos], default=0)
        fecha_fuente = datetime.utcnow()

        rubros_actualizados = 0
        for fila in filas_excel:
            if fila["tipo"] != "RUBRO" or fila["item"] not in safe_items:
                continue
            nodo = nodos_por_item.get(fila["item"])
            if not nodo:
                continue
            nodo.descripcion = fila["descripcion"]
            nodo.unidad = fila["unidad"]
            nodo.metrado = fila["metrado"]
            nodo.precio_unitario_ref = fila["precio_unitario_ref"]
            nodo.estado_actualizacion = "activo"
            nodo.actualizacion_lote_id = lote.id
            nodo.excel_fila = fila["row"]
            nodo.excel_hoja = hoja
            nodo.excel_archivo = archivo.filename
            nodo.fecha_actualizacion_fuente = fecha_fuente
            rubros_actualizados += 1

        obsoletos_marcados = 0
        for item in obsolete_items:
            nodo = nodos_por_item.get(item)
            if not nodo:
                continue
            nodo.estado_actualizacion = "obsoleto"
            nodo.actualizacion_lote_id = lote.id
            nodo.fecha_actualizacion_fuente = fecha_fuente
            obsoletos_marcados += 1

        stack_por_nivel: dict[int, NodoPresupuesto] = {}
        nodos_creados = 0
        for fila in filas_excel:
            nivel = ORDEN_JERARQUIA.index(fila["tipo"])
            item = fila["item"]
            nodo_existente = nodos_por_item.get(item) if item else None

            if nodo_existente is not None:
                nodo_existente.estado_actualizacion = "activo"
                stack_por_nivel[nivel] = nodo_existente
                for nivel_borrar in [n for n in stack_por_nivel if n > nivel]:
                    del stack_por_nivel[nivel_borrar]
                continue

            if not item or item not in items_a_crear:
                continue

            padre = None
            for nivel_padre in range(nivel - 1, -1, -1):
                if nivel_padre in stack_por_nivel:
                    padre = stack_por_nivel[nivel_padre]
                    break

            max_orden += 1
            es_rubro = fila["tipo"] == "RUBRO"
            nuevo = NodoPresupuesto(
                proyecto_id=proyecto_id,
                padre_id=padre.id if padre else None,
                tipo=fila["tipo"],
                nivel=nivel,
                item=item,
                descripcion=fila["descripcion"],
                orden=max_orden,
                unidad=fila["unidad"] if es_rubro else None,
                metrado=fila["metrado"] if es_rubro else None,
                precio_unitario_ref=fila["precio_unitario_ref"] if es_rubro else None,
                activo_como_rubro=es_rubro,
                tipo_rubro="PENDIENTE" if es_rubro else None,
                observaciones=fila["observaciones_excel"] if es_rubro else None,
                estado_actualizacion="activo",
                actualizacion_lote_id=lote.id,
                excel_fila=fila["row"],
                excel_hoja=hoja,
                excel_archivo=archivo.filename,
                fecha_actualizacion_fuente=fecha_fuente,
            )
            db.add(nuevo)
            db.flush()
            nodos_por_item[item] = nuevo
            stack_por_nivel[nivel] = nuevo
            for nivel_borrar in [n for n in stack_por_nivel if n > nivel]:
                del stack_por_nivel[nivel_borrar]
            nodos_creados += 1

        lote.total_nodos_creados = nodos_creados
        lote.total_rubros_actualizados = rubros_actualizados
        lote.total_obsoletos_marcados = obsoletos_marcados
        lote.resumen_json = json.dumps(
            {
                **resumen,
                "aplicado": {
                    "lote_id": lote.id,
                    "nodos_creados": nodos_creados,
                    "rubros_actualizados": rubros_actualizados,
                    "obsoletos_marcados": obsoletos_marcados,
                    "excepciones_no_aplicadas": resumen["rubros_revision"],
                },
            },
            ensure_ascii=False,
        )
        proyecto.fecha_actualizacion = datetime.utcnow()
        db.commit()

        return {
            "mensaje": "Actualizacion aplicada",
            "lote_id": lote.id,
            "resumen": {
                **resumen,
                "nodos_creados": nodos_creados,
                "rubros_actualizados": rubros_actualizados,
                "obsoletos_marcados": obsoletos_marcados,
                "excepciones_no_aplicadas": resumen["rubros_revision"],
            },
        }
    except Exception:
        db.rollback()
        raise


@router.patch("/nodos/{nodo_id}/vincular-apu", response_model=NodoOut)
def vincular_apu(
    nodo_id: int, data: VincularAPURequest, db: Session = Depends(get_db)
):
    """
    Vincula un APU a un nodo RUBRO.
    La unidad del rubro queda como referencia visual; no bloquea la vinculacion.
    """
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(
            status_code=400, detail="Solo se puede vincular un APU a nodos operativos tipo rubro"
        )

    apu = db.query(APU).filter(APU.id == data.apu_id).first()
    if not apu:
        raise HTTPException(status_code=404, detail="APU no encontrado")

    nodo.apu_id = data.apu_id
    nodo.tipo_rubro = "VINCULADO"
    nodo.observaciones = None
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)


@router.get("/nodos/{nodo_id}/revision-apu", response_model=RevisionAPUOut)
def obtener_revision_apu(nodo_id: int, db: Session = Depends(get_db)):
    nodo = (
        db.query(NodoPresupuesto)
        .options(joinedload(NodoPresupuesto.apu).joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(NodoPresupuesto.id == nodo_id)
        .first()
    )
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se puede revisar un rubro operativo")
    if not nodo.apu_id:
        raise HTTPException(status_code=400, detail="El rubro no tiene APU vinculado")
    return _revision_apu_out(nodo, db)


@router.post("/nodos/{nodo_id}/revision-apu/revisado", response_model=RevisionAPUOut)
def marcar_revision_apu_revisada(
    nodo_id: int,
    data: RevisionAPURevisadoIn,
    db: Session = Depends(get_db),
):
    nodo = (
        db.query(NodoPresupuesto)
        .options(joinedload(NodoPresupuesto.apu).joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(NodoPresupuesto.id == nodo_id)
        .first()
    )
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se puede revisar un rubro operativo")
    if not nodo.apu_id:
        raise HTTPException(status_code=400, detail="El rubro no tiene APU vinculado")

    costo_apu = _costo_apu_con_control(nodo.apu)
    motivos = _motivos_revision_apu(nodo, costo_apu)
    if not motivos:
        raise HTTPException(status_code=400, detail="El rubro no tiene motivos de revision pendientes")

    motivos_por_codigo = {motivo["codigo"]: motivo for motivo in motivos}
    items_por_codigo = {item.codigo_motivo: item for item in data.items}
    faltantes = [codigo for codigo in motivos_por_codigo if codigo not in items_por_codigo]
    extras = [codigo for codigo in items_por_codigo if codigo not in motivos_por_codigo]
    no_aprobados = [item.codigo_motivo for item in data.items if item.codigo_motivo in motivos_por_codigo and not item.aprobado]
    if faltantes or extras or no_aprobados:
        raise HTTPException(
            status_code=400,
            detail={
                "mensaje": "Para marcar como revisado debes aprobar todos los motivos actuales.",
                "faltantes": faltantes,
                "extras": extras,
                "no_aprobados": no_aprobados,
            },
        )

    firma = _firma_revision_apu(nodo, motivos)
    revision = NodoAPURevision(
        nodo_id=nodo.id,
        apu_id=nodo.apu_id,
        estado="validado",
        firma_revision=firma,
        snapshot_descripcion=nodo.descripcion or "",
        snapshot_unidad=nodo.unidad or "",
        snapshot_apu_id=nodo.apu_id,
    )
    revision.items = [
        NodoAPURevisionItem(
            codigo_motivo=motivo["codigo"],
            descripcion_motivo=motivo["descripcion"],
            aprobado=True,
            comentario=(items_por_codigo[motivo["codigo"]].comentario or "").strip() or None,
        )
        for motivo in motivos
    ]
    db.add(revision)
    db.commit()
    db.refresh(nodo)
    return _revision_apu_out(nodo, db, costo_apu)


@router.patch("/nodos/{nodo_id}/desvincular-apu", response_model=NodoOut)
def desvincular_apu(nodo_id: int, db: Session = Depends(get_db)):
    """
    Quita el APU vinculado a un nodo RUBRO.
    El nodo vuelve a estado PENDIENTE.
    """
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")

    nodo.apu_id = None
    nodo.tipo_rubro = "PENDIENTE"
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)


@router.post("/nodos/{nodo_id}/crear-apu", response_model=CrearAPUDesdeRubroOut, status_code=201)
def crear_apu_desde_rubro(
    nodo_id: int,
    data: Optional[CrearAPUDesdeRubroRequest] = None,
    db: Session = Depends(get_db),
):
    """
    Crea un APU en revision desde un RUBRO y lo vincula en la misma transaccion.
    """
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se puede crear APU desde nodos operativos tipo rubro")
    if not nodo.unidad:
        raise HTTPException(status_code=400, detail="El rubro no tiene unidad para crear el APU")

    base_apu = None
    if data and data.base_apu_id:
        base_apu = (
            db.query(APU)
            .filter(APU.id == data.base_apu_id)
            .first()
        )
        if not base_apu:
            raise HTTPException(status_code=404, detail="APU base no encontrado")

    apu = APU(
        codigo=siguiente_codigo_apu(db),
        nombre=nodo.descripcion,
        descripcion=nodo.descripcion,
        unidad=nodo.unidad,
        categoria=base_apu.categoria if base_apu else None,
        subcategoria=base_apu.subcategoria if base_apu else None,
        rendimiento=base_apu.rendimiento if base_apu else 1.0,
        estado="en_revision",
        observacion=f"Duplicado desde {base_apu.codigo or base_apu.nombre}" if base_apu else None,
    )
    db.add(apu)
    db.flush()

    if base_apu:
        base_items = db.query(APUItem).filter(APUItem.apu_id == base_apu.id).all()
        for item in base_items:
            db.add(
                APUItem(
                    apu_id=apu.id,
                    recurso_id=item.recurso_id,
                    categoria=item.categoria,
                    cantidad=item.cantidad,
                    orden=item.orden,
                    es_herramienta_menor=item.es_herramienta_menor,
                )
            )

    nodo.apu_id = apu.id
    nodo.tipo_rubro = "VINCULADO"
    nodo.observaciones = None

    db.commit()
    db.refresh(nodo)
    db.refresh(apu)
    return {
        "nodo": nodo,
        "apu_id": apu.id,
        "codigo": apu.codigo,
        "nombre": apu.nombre,
        "unidad": apu.unidad,
        "estado": apu.estado,
    }


@router.get("/proyectos/{proyecto_id}/apus/{apu_base_id}/variantes", response_model=list[VarianteAPUOut])
def listar_variantes_apu(proyecto_id: int, apu_base_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    base = db.query(APU).filter(APU.id == apu_base_id).first()
    if not base:
        raise HTTPException(status_code=404, detail="APU base no encontrado")
    if base.es_variante:
        apu_base_id = _apu_base_id(base)

    variantes = (
        db.query(APU)
        .options(joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(
            APU.es_variante.is_(True),
            APU.proyecto_id == proyecto_id,
            APU.apu_base_id == apu_base_id,
        )
        .order_by(APU.variante_nombre, APU.id)
        .all()
    )
    return [_variante_out(db, variante) for variante in variantes]


@router.post("/proyectos/{proyecto_id}/apus/{apu_id}/aislar-no-liberados", response_model=AislarAPUNoLiberadosOut)
def aislar_apu_no_liberados(proyecto_id: int, apu_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto.id).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    source = (
        db.query(APU)
        .options(joinedload(APU.items))
        .filter(APU.id == apu_id)
        .first()
    )
    if not source:
        raise HTTPException(status_code=404, detail="APU no encontrado")

    base_apu = _apu_base(db, source)
    nodos = db.query(NodoPresupuesto).filter(NodoPresupuesto.proyecto_id == proyecto_id).all()
    hijos_por_padre: dict[Optional[int], list[NodoPresupuesto]] = {}
    for nodo in nodos:
        hijos_por_padre.setdefault(nodo.padre_id, []).append(nodo)

    ids_liberados: set[int] = set()
    paquetes_liberados = (
        db.query(PaquetePresupuesto)
        .filter(PaquetePresupuesto.proyecto_id == proyecto_id, PaquetePresupuesto.estado == "liberado")
        .all()
    )
    for paquete in paquetes_liberados:
        ids_liberados.add(paquete.nodo_id)
        ids_liberados.update(_descendientes_ids(paquete.nodo_id, hijos_por_padre))

    usos = [nodo for nodo in nodos if nodo.apu_id == apu_id]
    rubros_liberados = [nodo for nodo in usos if nodo.id in ids_liberados]
    rubros_no_liberados = [nodo for nodo in usos if nodo.id not in ids_liberados]
    if not rubros_no_liberados:
        raise HTTPException(status_code=400, detail="No hay rubros no liberados para aislar")

    variante_nombre = "NO LIBERADOS"
    variante = _buscar_variante_existente(db, proyecto_id, base_apu.id, variante_nombre)
    created = False
    if not variante:
        variante = APU(
            codigo=_siguiente_codigo_variante(db, base_apu),
            nombre=f"{base_apu.nombre} - {variante_nombre}",
            descripcion=source.descripcion or base_apu.descripcion,
            categoria=source.categoria,
            subcategoria=source.subcategoria,
            unidad=source.unidad,
            rendimiento=source.rendimiento,
            estado="en_revision",
            version=1,
            observacion=f"Variante '{variante_nombre}' de {base_apu.codigo or base_apu.nombre}",
            es_variante=True,
            apu_base_id=base_apu.id,
            proyecto_id=proyecto_id,
            variante_nombre=variante_nombre,
            copiado_desde_apu_id=source.id,
        )
        db.add(variante)
        db.flush()
        _copiar_items_apu(db, source, variante)
        created = True

    for nodo in rubros_no_liberados:
        nodo.apu_id = variante.id
        nodo.tipo_rubro = "VINCULADO"
        nodo.observaciones = None

    db.commit()
    db.refresh(variante)
    return AislarAPUNoLiberadosOut(
        variante=_variante_out(db, variante),
        apu_id=variante.id,
        rubros_reasignados=len(rubros_no_liberados),
        rubros_liberados_preservados=len(rubros_liberados),
        created=created,
    )


@router.post("/proyectos/{proyecto_id}/apus/{apu_id}/recursos/{recurso_id}/copiar-para-proyecto", response_model=RecursoProyectoOut)
def copiar_recurso_para_proyecto(
    proyecto_id: int,
    apu_id: int,
    recurso_id: int,
    data: RecursoProyectoUpdate,
    db: Session = Depends(get_db),
):
    proyecto = db.query(Proyecto.id).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    apu = db.query(APU).filter(APU.id == apu_id).first()
    if not apu:
        raise HTTPException(status_code=404, detail="APU no encontrado")
    if not apu.es_variante or apu.proyecto_id != proyecto_id:
        raise HTTPException(status_code=400, detail="Primero aisla o crea una variante APU para este proyecto")
    recurso_base = db.query(Recurso).filter(Recurso.id == recurso_id).first()
    if not recurso_base:
        raise HTTPException(status_code=404, detail="Recurso no encontrado")
    item_existe = (
        db.query(APUItem.id)
        .filter(APUItem.apu_id == apu_id, APUItem.recurso_id == recurso_id)
        .first()
    )
    if not item_existe:
        raise HTTPException(status_code=404, detail="El recurso no esta usado por este APU")
    if data.precio_unitario < 0 or not math.isfinite(data.precio_unitario):
        raise HTTPException(status_code=400, detail="Precio invalido")
    descripcion = " ".join(data.descripcion.split())
    unidad = " ".join(data.unidad.split())
    if not descripcion or not unidad:
        raise HTTPException(status_code=400, detail="Nombre y unidad del recurso son obligatorios")

    recurso = (
        db.query(Recurso)
        .filter(Recurso.proyecto_id == proyecto_id, Recurso.recurso_base_id == recurso_id)
        .first()
    )
    created = False
    if not recurso:
        recurso = Recurso(
            proyecto_id=proyecto_id,
            recurso_base_id=recurso_id,
            codigo=_codigo_recurso_proyecto(db, recurso_base, proyecto_id),
            descripcion=descripcion,
            categoria=recurso_base.categoria,
            subcategoria=recurso_base.subcategoria,
            familia=recurso_base.familia,
            unidad=unidad,
            precio_unitario=data.precio_unitario,
            fecha_precio=recurso_base.fecha_precio,
            fecha_validacion=recurso_base.fecha_validacion,
            etiquetas=list(recurso_base.etiquetas or []),
            activo=True,
        )
        db.add(recurso)
        db.flush()
        created = True

    recurso.descripcion = descripcion
    recurso.unidad = unidad
    recurso.precio_unitario = data.precio_unitario
    recurso.fuente_precio = data.fuente_precio
    recurso.observacion = data.observacion
    recurso.estado_validacion = data.estado_validacion or "pendiente"
    recurso.fuente_validacion = "PRESUPUESTOS"
    recurso.nota_validacion = data.nota_validacion
    if "especial del proyecto" not in (recurso.etiquetas or []):
        recurso.etiquetas = [*(recurso.etiquetas or []), "especial del proyecto"]

    (
        db.query(APUItem)
        .filter(APUItem.apu_id == apu_id, APUItem.recurso_id == recurso_id)
        .update({APUItem.recurso_id: recurso.id}, synchronize_session=False)
    )
    db.commit()
    db.refresh(recurso)
    return RecursoProyectoOut(recurso=recurso, recurso_id=recurso.id, apu_id=apu_id, created=created)


@router.post("/nodos/{nodo_id}/variantes-apu", response_model=VarianteAPUAsignadaOut, status_code=201)
def crear_variante_apu(
    nodo_id: int,
    data: CrearVarianteAPURequest,
    db: Session = Depends(get_db),
):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se puede crear variante en nodos operativos tipo rubro")
    if not nodo.apu_id:
        raise HTTPException(status_code=400, detail="El rubro no tiene APU vinculado")

    apu_actual = db.query(APU).filter(APU.id == nodo.apu_id).first()
    if not apu_actual:
        raise HTTPException(status_code=404, detail="APU vinculado no encontrado")
    base_apu = _apu_base(db, apu_actual)
    variante_nombre = _normalizar_variante_nombre(data.variante_nombre)

    existente = _buscar_variante_existente(db, nodo.proyecto_id, base_apu.id, variante_nombre)
    if existente:
        nodo.apu_id = existente.id
        nodo.tipo_rubro = "VINCULADO"
        nodo.observaciones = None
        db.commit()
        db.refresh(nodo)
        return {
            "nodo": _nodo_out(nodo),
            "variante": _variante_out(db, existente),
            "apu_id": existente.id,
            "created": False,
        }

    source_id = data.copiar_desde_apu_id or apu_actual.id
    source = _validar_fuente_variante(db, source_id, nodo.proyecto_id, base_apu.id)

    variante = APU(
        codigo=_siguiente_codigo_variante(db, base_apu),
        nombre=f"{base_apu.nombre} - {variante_nombre}",
        descripcion=source.descripcion or base_apu.descripcion,
        categoria=source.categoria,
        subcategoria=source.subcategoria,
        unidad=source.unidad,
        rendimiento=source.rendimiento,
        estado="en_revision",
        version=1,
        observacion=f"Variante '{variante_nombre}' de {base_apu.codigo or base_apu.nombre}",
        es_variante=True,
        apu_base_id=base_apu.id,
        proyecto_id=nodo.proyecto_id,
        variante_nombre=variante_nombre,
        copiado_desde_apu_id=source.id,
    )
    db.add(variante)
    db.flush()
    _copiar_items_apu(db, source, variante)

    nodo.apu_id = variante.id
    nodo.tipo_rubro = "VINCULADO"
    nodo.observaciones = None

    db.commit()
    db.refresh(nodo)
    db.refresh(variante)
    return {
        "nodo": _nodo_out(nodo),
        "variante": _variante_out(db, variante),
        "apu_id": variante.id,
        "created": True,
    }


@router.patch("/nodos/{nodo_id}/variante-apu", response_model=VarianteAPUAsignadaOut)
def cambiar_variante_apu(
    nodo_id: int,
    data: CambiarVarianteAPURequest,
    db: Session = Depends(get_db),
):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se puede cambiar variante en nodos operativos tipo rubro")
    if not nodo.apu_id:
        raise HTTPException(status_code=400, detail="El rubro no tiene APU vinculado")

    apu_actual = db.query(APU).filter(APU.id == nodo.apu_id).first()
    if not apu_actual:
        raise HTTPException(status_code=404, detail="APU vinculado no encontrado")
    base_apu = _apu_base(db, apu_actual)

    if data.variante_apu_id is None:
        nodo.apu_id = base_apu.id
        nodo.tipo_rubro = "VINCULADO"
        nodo.observaciones = None
        db.commit()
        db.refresh(nodo)
        return {
            "nodo": _nodo_out(nodo),
            "variante": None,
            "apu_id": base_apu.id,
            "created": False,
        }

    variante = (
        db.query(APU)
        .options(joinedload(APU.items).joinedload(APUItem.recurso))
        .filter(APU.id == data.variante_apu_id)
        .first()
    )
    if not variante:
        raise HTTPException(status_code=404, detail="Variante no encontrada")
    if not variante.es_variante or variante.proyecto_id != nodo.proyecto_id or variante.apu_base_id != base_apu.id:
        raise HTTPException(status_code=400, detail="La variante no corresponde al APU base del rubro")

    nodo.apu_id = variante.id
    nodo.tipo_rubro = "VINCULADO"
    nodo.observaciones = None
    db.commit()
    db.refresh(nodo)
    return {
        "nodo": _nodo_out(nodo),
        "variante": _variante_out(db, variante),
        "apu_id": variante.id,
        "created": False,
    }

@router.patch("/nodos/{nodo_id}/individualizar", response_model=NodoOut)
def individualizar_nodo(nodo_id: int, db: Session = Depends(get_db)):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    nodo.individualizado = True
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)


@router.patch("/nodos/{nodo_id}/reagrupar", response_model=NodoOut)
def reagrupar_nodo(nodo_id: int, db: Session = Depends(get_db)):
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    nodo.individualizado = False
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)

@router.patch("/nodos/{nodo_id}/marcar-sin-apu", response_model=NodoOut)
def marcar_sin_apu(nodo_id: int, db: Session = Depends(get_db)):
    """Marca un rubro como Sin APU manualmente."""
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not _es_rubro_operativo(db, nodo):
        raise HTTPException(status_code=400, detail="Solo se pueden marcar nodos operativos tipo rubro")
    # Si tenía APU vinculado, lo quitamos
    nodo.apu_id = None
    nodo.tipo_rubro = "PENDIENTE"
    nodo.observaciones = "SIN_APU"
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)


@router.patch("/nodos/{nodo_id}/desmarcar-sin-apu", response_model=NodoOut)
def desmarcar_sin_apu(nodo_id: int, db: Session = Depends(get_db)):
    """Quita la marca Sin APU — el rubro vuelve a estado Pendiente."""
    nodo = db.query(NodoPresupuesto).filter(NodoPresupuesto.id == nodo_id).first()
    if not nodo:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    nodo.observaciones = None
    nodo.tipo_rubro = "PENDIENTE"
    db.commit()
    db.refresh(nodo)
    return _nodo_out(nodo)
